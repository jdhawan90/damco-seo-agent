"""
Vendor Scorer — Phase 2 module of Off-Page & Links
===================================================

Computes platform / vendor performance from historical
`offpage_activities` and rolls the result back into `platform_targets`.
Surfaces which platforms are worth continued outreach and which are
exhausted (or near-spammy).

What gets computed (per platform_target)
----------------------------------------
- attempts            — total drafts + outreach + guest posts sent
- responses           — count of activities with status != 'no_response'
                        (anything that progressed past 'submitted')
- publications        — count with status='published'
- rejections          — count with status='rejected'
- no_responses        — count with status='no_response'
- still_draft         — count with status='draft' (not yet sent)
- response_rate       — responses / attempts (0-100%)
- publication_rate    — publications / attempts (0-100%)
- avg_turnaround_days — mean days from submit → publish
- avg_da_link         — average DA of pages that linked back (joined to backlinks)
- last_contacted      — max activity.date for the platform
- quality_score       — 0-100 weighted aggregate:
                        pub_rate*0.5 + resp_rate*0.25 + da_score*0.15 + recency_score*0.10

Status mutations
----------------
- If response_rate < 10% AND attempts >= 5 → mark as 'exhausted'
  (don't draft to them again until manually reset)
- If publications > 0 → keep 'active' regardless of low response rate
- Platforms with attempts == 0 → unchanged (we haven't really tried)

Outputs
-------
- `platform_targets` updates: response_rate, quality_score, last_contacted,
  status (when transitioning to 'exhausted')
- `outputs/audits/vendor_scores_<date>.md` — sortable narrative + recommendations
- `outputs/reports/vendor_scores_<date>.xlsx` — sortable data

Usage
-----
    # Default — score every platform that has any activity history
    python -m offpage_links.vendor_scorer

    # Restrict to active platforms (skip blacklist + already-exhausted)
    python -m offpage_links.vendor_scorer --only-active

    # Tune the auto-exhaustion threshold (default response_rate < 10%)
    python -m offpage_links.vendor_scorer --exhaust-below 15

    # Dry run — analysis + report, no DB writes / no status mutations
    python -m offpage_links.vendor_scorer --dry-run

Design notes
------------
Pure SQL aggregation + Python tabulation. No LLM, no API calls.
"""

from __future__ import annotations

import argparse
import logging
import sys
import time
from datetime import date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from common.database import connection, fetch_all, record_agent_run


logger = logging.getLogger("vendor_scorer")
AGENT_NAME = "offpage_links.vendor_scorer"

OUTPUT_AUDITS  = Path(__file__).resolve().parent.parent / "outputs" / "audits"
OUTPUT_REPORTS = Path(__file__).resolve().parent.parent / "outputs" / "reports"

DEFAULT_EXHAUST_BELOW_PCT = 10.0
MIN_ATTEMPTS_FOR_EXHAUST = 5
RECENCY_DECAY_DAYS = 180         # links older than this don't contribute to recency_score


# ---------------------------------------------------------------------------
# Read phase
# ---------------------------------------------------------------------------

def load_activity_aggregates(only_active: bool) -> list[dict]:
    """
    For each platform, return raw counts. We compute response_rate /
    quality_score in Python so the math is auditable.
    """
    sql = """
        SELECT pt.id                                   AS platform_id,
               pt.platform_url,
               pt.platform_name,
               pt.domain_authority                     AS pt_da,
               pt.status                               AS pt_status,
               pt.niche,
               pt.last_contacted                       AS pt_last_contacted,
               count(oa.*)                             AS attempts,
               count(*) FILTER (WHERE oa.status NOT IN ('draft', 'no_response')) AS responses,
               count(*) FILTER (WHERE oa.status = 'published')   AS publications,
               count(*) FILTER (WHERE oa.status = 'rejected')    AS rejections,
               count(*) FILTER (WHERE oa.status = 'no_response') AS no_responses,
               count(*) FILTER (WHERE oa.status = 'draft')       AS still_draft,
               count(*) FILTER (WHERE oa.status = 'submitted')   AS submitted,
               max(oa.date)                            AS last_activity,
               min(oa.date)                            AS first_activity
          FROM platform_targets pt
     LEFT JOIN offpage_activities oa ON oa.platform_id = pt.id
    """
    if only_active:
        sql += " WHERE pt.status = 'active'\n"
    sql += " GROUP BY pt.id, pt.platform_url, pt.platform_name, pt.domain_authority, pt.status, pt.niche, pt.last_contacted"
    return fetch_all(sql)


def load_turnaround_per_platform() -> dict[int, float]:
    """
    Average days from earliest 'submitted' to 'published' per platform.
    Approximation — we don't have per-activity status-transition timestamps,
    so we approximate via (publication date - earliest activity date) per
    platform, which is close enough for vendor scoring.
    """
    rows = fetch_all(
        """
        WITH per_platform AS (
            SELECT platform_id,
                   min(CASE WHEN status = 'submitted' THEN date END) AS first_submit,
                   min(CASE WHEN status = 'published' THEN date END) AS first_publish
              FROM offpage_activities
             WHERE platform_id IS NOT NULL
             GROUP BY platform_id
        )
        SELECT platform_id, first_submit, first_publish,
               (first_publish - first_submit) AS days_to_publish
          FROM per_platform
         WHERE first_submit IS NOT NULL AND first_publish IS NOT NULL
        """
    )
    return {r["platform_id"]: float(r["days_to_publish"])
            for r in rows if r.get("days_to_publish") is not None}


def load_avg_da_per_platform() -> dict[int, float]:
    """
    For each platform, average DA of the backlinks they generated.
    Joined via: offpage_activities.target_page_id  ->  backlinks.page_id  ->  source_domain
    matched against platform's root domain.
    """
    rows = fetch_all(
        """
        SELECT oa.platform_id,
               avg(b.domain_authority)::float AS avg_da
          FROM offpage_activities oa
          JOIN backlinks b ON b.page_id = oa.target_page_id
         WHERE oa.platform_id IS NOT NULL
           AND b.source_domain IS NOT NULL
           AND b.domain_authority IS NOT NULL
           -- best-effort match: backlink's source_domain contains the platform name
           AND lower(b.source_domain) ILIKE '%' || lower(
                    regexp_replace(
                        coalesce((SELECT platform_url FROM platform_targets WHERE id = oa.platform_id), ''),
                        '^https?://(www\\.)?|/.*$', '', 'g')
               ) || '%'
         GROUP BY oa.platform_id
        """
    )
    return {r["platform_id"]: float(r["avg_da"])
            for r in rows if r.get("avg_da") is not None}


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------

def compute_scores(rows: list[dict],
                    turnaround: dict[int, float],
                    avg_da: dict[int, float]) -> list[dict]:
    out: list[dict] = []
    today = date.today()

    for r in rows:
        attempts = r["attempts"] or 0
        responses = r["responses"] or 0
        publications = r["publications"] or 0

        response_rate = (100.0 * responses / attempts) if attempts else 0.0
        publication_rate = (100.0 * publications / attempts) if attempts else 0.0

        last_activity = r["last_activity"]
        days_since_last = ((today - last_activity).days
                           if last_activity else None)

        # Recency score: 100 if linked in last 30d, decays linearly to 0 at RECENCY_DECAY_DAYS
        if days_since_last is None:
            recency_score = 0.0
        elif days_since_last <= 30:
            recency_score = 100.0
        elif days_since_last >= RECENCY_DECAY_DAYS:
            recency_score = 0.0
        else:
            recency_score = 100.0 * (1.0 - (days_since_last - 30) / (RECENCY_DECAY_DAYS - 30))

        # DA score: scale 0-100 DA to 0-100 directly (clamp)
        platform_da = avg_da.get(r["platform_id"]) or (float(r["pt_da"]) if r["pt_da"] else 0.0)
        da_score = max(0.0, min(100.0, platform_da))

        quality_score = round(
            publication_rate * 0.50
            + response_rate   * 0.25
            + da_score        * 0.15
            + recency_score   * 0.10,
            2,
        )

        out.append({
            "platform_id":         r["platform_id"],
            "platform_url":        r["platform_url"],
            "platform_name":       r["platform_name"],
            "current_status":      r["pt_status"],
            "niche":               r["niche"],
            "attempts":            attempts,
            "responses":           responses,
            "publications":        publications,
            "rejections":          r["rejections"] or 0,
            "no_responses":        r["no_responses"] or 0,
            "still_draft":         r["still_draft"] or 0,
            "response_rate":       round(response_rate, 2),
            "publication_rate":    round(publication_rate, 2),
            "avg_turnaround_days": round(turnaround.get(r["platform_id"], 0.0), 1)
                                    if turnaround.get(r["platform_id"]) else None,
            "platform_da":         round(platform_da, 1) if platform_da else None,
            "last_activity":       last_activity.isoformat() if last_activity else None,
            "days_since_last":     days_since_last,
            "recency_score":       round(recency_score, 1),
            "quality_score":       quality_score,
        })

    out.sort(key=lambda x: -x["quality_score"])
    return out


def decide_status_updates(scored: list[dict],
                           exhaust_below_pct: float) -> list[dict]:
    """
    Returns a list of dicts {platform_id, new_status, reason} for rows that
    should transition. Does not mutate the input.
    """
    updates: list[dict] = []
    for s in scored:
        cur = s["current_status"]
        # Already exhausted / blacklisted — never auto-resurrect
        if cur in ("blacklist", "exhausted"):
            continue

        # Auto-exhaust low responders with non-trivial attempts and NO publications
        if (s["response_rate"] < exhaust_below_pct
                and s["attempts"] >= MIN_ATTEMPTS_FOR_EXHAUST
                and s["publications"] == 0):
            updates.append({
                "platform_id": s["platform_id"],
                "new_status":  "exhausted",
                "reason":      (f"{s['response_rate']:.1f}% response rate over "
                                f"{s['attempts']} attempts and 0 publications "
                                f"(threshold: <{exhaust_below_pct}%)"),
            })
    return updates


# ---------------------------------------------------------------------------
# Write phase
# ---------------------------------------------------------------------------

def apply_scores_to_platform_targets(scored: list[dict],
                                      status_updates: list[dict]) -> tuple[int, int]:
    """
    Update platform_targets with refreshed response_rate, quality_score,
    last_contacted, status. Returns (rows_updated, status_changed).
    """
    by_id_status = {u["platform_id"]: u for u in status_updates}
    rows_updated = 0
    status_changed = 0
    with connection() as conn:
        with conn.cursor() as cur:
            for s in scored:
                update = by_id_status.get(s["platform_id"])
                if update:
                    cur.execute(
                        """
                        UPDATE platform_targets
                           SET response_rate  = %s,
                               quality_score  = %s,
                               last_contacted = %s,
                               status         = %s
                         WHERE id = %s
                        """,
                        (s["response_rate"], s["quality_score"],
                         s["last_activity"], update["new_status"], s["platform_id"]),
                    )
                    status_changed += 1
                else:
                    cur.execute(
                        """
                        UPDATE platform_targets
                           SET response_rate  = %s,
                               quality_score  = %s,
                               last_contacted = %s
                         WHERE id = %s
                        """,
                        (s["response_rate"], s["quality_score"],
                         s["last_activity"], s["platform_id"]),
                    )
                rows_updated += cur.rowcount
    return rows_updated, status_changed


def write_markdown(scored: list[dict], status_updates: list[dict],
                    exhaust_below_pct: float, only_active: bool) -> Path:
    OUTPUT_AUDITS.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_AUDITS / f"vendor_scores_{date.today().isoformat()}.md"

    p: list[str] = []
    p.append(f"# Vendor / Platform Performance — {date.today().isoformat()}")
    p.append("")
    p.append(f"_Generated by `{AGENT_NAME}`._")
    if only_active:
        p.append("_Scope: only `status='active'` platforms_")
    p.append("")

    p.append("## Summary")
    p.append("")
    p.append("| Metric | Value |")
    p.append("|---|---:|")
    p.append(f"| Platforms analyzed | {len(scored)} |")
    p.append(f"| Platforms with ≥1 activity | {sum(1 for s in scored if s['attempts'] > 0)} |")
    p.append(f"| Total publications | {sum(s['publications'] for s in scored)} |")
    p.append(f"| Total attempts | {sum(s['attempts'] for s in scored)} |")
    p.append(f"| Platforms recommended → `exhausted` | {len(status_updates)} |")
    p.append(f"| Exhaust threshold | response_rate < {exhaust_below_pct}% "
             f"with ≥{MIN_ATTEMPTS_FOR_EXHAUST} attempts and 0 publications |")
    p.append("")

    # Top performers (where we should keep investing)
    productive = [s for s in scored if s["attempts"] > 0]
    if productive:
        p.append("## Top performers (quality score)")
        p.append("")
        p.append("| # | Platform | Quality | Attempts | Pub | Resp Rate | Pub Rate | DA | Last Act. |")
        p.append("|---:|---|---:|---:|---:|---:|---:|---:|---|")
        for i, s in enumerate(productive[:15], 1):
            da = f"{s['platform_da']}" if s['platform_da'] else "—"
            last = s["last_activity"] or "—"
            p.append(f"| {i} | `{s['platform_url']}` | {s['quality_score']} | "
                     f"{s['attempts']} | {s['publications']} | "
                     f"{s['response_rate']}% | {s['publication_rate']}% | "
                     f"{da} | {last} |")
        p.append("")

    # Status mutations recommended
    if status_updates:
        p.append("## ⚠️ Status changes applied (or staged in --dry-run)")
        p.append("")
        for u in status_updates:
            scored_row = next((s for s in scored if s["platform_id"] == u["platform_id"]), None)
            url = scored_row["platform_url"] if scored_row else u["platform_id"]
            p.append(f"- `{url}` → **{u['new_status']}** — {u['reason']}")
        p.append("")

    # Underperformers worth retiring  (not yet auto-exhausted but trending poorly)
    near_exhaust = [s for s in scored
                    if s["attempts"] >= 3 and s["publications"] == 0
                    and s not in [next((x for x in scored if x["platform_id"] == u["platform_id"]), None)
                                  for u in status_updates]]
    if near_exhaust:
        p.append("## ⚠️ Approaching the exhaustion threshold")
        p.append("")
        p.append("| Platform | Attempts | Pub | Resp Rate | Last Act. |")
        p.append("|---|---:|---:|---:|---|")
        for s in near_exhaust[:10]:
            last = s["last_activity"] or "—"
            p.append(f"| `{s['platform_url']}` | {s['attempts']} | {s['publications']} | "
                     f"{s['response_rate']}% | {last} |")
        p.append("")

    # Inactive  (no activity at all)
    inactive = [s for s in scored if s["attempts"] == 0]
    if inactive:
        p.append("## Inactive platforms (no outreach attempted yet)")
        p.append("")
        p.append(f"_{len(inactive)} platform target(s) in the DB have had zero "
                 f"`offpage_activities` rows. Consider running "
                 f"`outreach_drafter` against the top scorers from "
                 f"`platform_finder`._")
        p.append("")

    path.write_text("\n".join(p), encoding="utf-8")
    return path


def write_excel(scored: list[dict]) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    OUTPUT_REPORTS.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_REPORTS / f"vendor_scores_{date.today().isoformat()}.xlsx"
    wb = openpyxl.Workbook()
    fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    bold = Font(bold=True, color="FFFFFF")

    ws = wb.active
    ws.title = "Vendor Scores"
    headers = [
        "Platform URL", "Name", "Current Status", "Niche",
        "Attempts", "Responses", "Publications", "Rejections",
        "No Response", "Still Draft", "Response Rate %", "Publication Rate %",
        "Turnaround (days)", "Platform DA", "Recency Score",
        "Quality Score", "Last Activity",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = bold
        c.fill = fill

    for s in scored:
        ws.append([
            s["platform_url"], s["platform_name"], s["current_status"], s["niche"] or "",
            s["attempts"], s["responses"], s["publications"], s["rejections"],
            s["no_responses"], s["still_draft"],
            s["response_rate"], s["publication_rate"],
            s["avg_turnaround_days"] or "",
            s["platform_da"] or "",
            s["recency_score"],
            s["quality_score"],
            s["last_activity"] or "",
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(only_active: bool = False,
        exhaust_below_pct: float = DEFAULT_EXHAUST_BELOW_PCT,
        dry_run: bool = False) -> dict:
    start = time.monotonic()

    aggregates = load_activity_aggregates(only_active)
    if not aggregates:
        logger.warning("No platform_targets rows%s.",
                       " with status='active'" if only_active else "")
        return {"status": "skipped", "reason": "no platforms"}

    turnaround = load_turnaround_per_platform()
    try:
        avg_da = load_avg_da_per_platform()
    except Exception as exc:
        # Backlinks join is best-effort; if the regex/cast fails (older psycopg2,
        # missing data), we just skip the DA boost rather than failing the whole run.
        logger.warning("avg-DA join failed (%s); proceeding without DA boost.", exc)
        avg_da = {}

    scored = compute_scores(aggregates, turnaround, avg_da)
    logger.info("Scored %d platform(s)", len(scored))

    status_updates = decide_status_updates(scored, exhaust_below_pct)
    rows_updated, status_changed = 0, 0
    if not dry_run:
        rows_updated, status_changed = apply_scores_to_platform_targets(scored, status_updates)
        logger.info("platform_targets updated: %d rows (status changes: %d)",
                    rows_updated, status_changed)

    md_path   = write_markdown(scored, status_updates, exhaust_below_pct, only_active)
    xlsx_path = write_excel(scored)

    duration = time.monotonic() - start

    if not dry_run:
        record_agent_run(
            agent_name=AGENT_NAME,
            status="success",
            records_processed=rows_updated,
            errors=[],
            duration_seconds=round(duration, 2),
            metadata={
                "run_date":            date.today().isoformat(),
                "only_active":         only_active,
                "exhaust_below_pct":   exhaust_below_pct,
                "platforms_scored":    len(scored),
                "platforms_with_activity": sum(1 for s in scored if s["attempts"] > 0),
                "total_publications":  sum(s["publications"] for s in scored),
                "rows_updated":        rows_updated,
                "status_changed":      status_changed,
                "md_path":             str(md_path),
                "xlsx_path":           str(xlsx_path),
            },
        )

    # Console summary
    print()
    print(f"  {'=' * 72}")
    print(f"   VENDOR SCORER -- {date.today().isoformat()}")
    print(f"  {'=' * 72}")
    print()
    print(f"  Platforms scored:        {len(scored)}")
    print(f"  Platforms with activity: {sum(1 for s in scored if s['attempts'] > 0)}")
    print(f"  Total publications:      {sum(s['publications'] for s in scored)}")
    print(f"  Status changes:          {status_changed}{' (dry-run)' if dry_run else ''}")
    print(f"  Exhaust threshold:       response_rate < {exhaust_below_pct}%")
    print(f"  Markdown:                {md_path}")
    print(f"  Excel:                   {xlsx_path}")
    print(f"  Duration:                {duration:.2f}s")
    productive = [s for s in scored if s["attempts"] > 0][:5]
    if productive:
        print()
        print("  Top 5 by quality score:")
        for s in productive:
            print(f"    {s['quality_score']:>6.1f}  {s['platform_url']:<40}  "
                  f"({s['publications']} pub / {s['attempts']} attempts)")
    print()

    return {
        "status":           "success",
        "scored":           len(scored),
        "rows_updated":     rows_updated,
        "status_changed":   status_changed,
        "md_path":          str(md_path),
        "xlsx_path":        str(xlsx_path),
        "duration_seconds": round(duration, 2),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Damco Outreach Vendor Scorer")
    parser.add_argument("--only-active", action="store_true",
                        help="Restrict to platforms with status='active'")
    parser.add_argument("--exhaust-below", type=float, default=DEFAULT_EXHAUST_BELOW_PCT,
                        help=f"Response-rate threshold below which we auto-exhaust "
                             f"(default: {DEFAULT_EXHAUST_BELOW_PCT}%%)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Compute + report but skip DB updates")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(name)s  %(levelname)s  %(message)s",
    )

    run(only_active=args.only_active,
        exhaust_below_pct=args.exhaust_below,
        dry_run=args.dry_run)


if __name__ == "__main__":
    main()
