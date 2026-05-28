"""
Gap Analyzer — first module of the Competitive Intelligence agent
==================================================================

For each tracked offering, surfaces where competitors rank in our SERPs
and Damco doesn't. Three gap types:

  coverage_gap  — Damco not in top 100; >= 1 tracked competitor in top 10
  displacement  — Damco at #11-30; competitor in top 10 (we have a page,
                  just need to outrank specific competitors)
  cluster_win   — Same competitor wins top-10 placements for >= 3 keywords
                  in the offering (they own a sub-niche)

Severity is GSC-traffic-weighted: gaps on keywords with real impressions
or clicks score higher than purely theoretical ones.

Outputs:
  outputs/reports/gap_analysis_<date>.xlsx  — multi-sheet matrix (one
      row per active keyword across all offerings)
  outputs/audits/gap_analysis_<offering>_<date>.md — per-offering
      narrative report. LLM-generated executive summary +
      recommendations when --with-narrative AND ANTHROPIC_API_KEY is set;
      rule-based fallback otherwise.

Usage
-----
    # All offerings, no LLM narrative, Excel + markdown
    python -m competitive_intelligence.gap_analyzer

    # One offering
    python -m competitive_intelligence.gap_analyzer --offering "AI"

    # Include LLM-generated narrative (needs ANTHROPIC_API_KEY + credit)
    python -m competitive_intelligence.gap_analyzer --offering "AI" --with-narrative

    # Generate report without DB writes / agent_runs row
    python -m competitive_intelligence.gap_analyzer --dry-run
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import time
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from common.database import fetch_all, record_agent_run
from common.llm import call_claude, LLMUnavailableError


logger = logging.getLogger("gap_analyzer")
AGENT_NAME = "competitive_intelligence.gap_analyzer"

# Detection thresholds
DISPLACEMENT_RANGE = (11, 30)   # Damco rank in this band = "displacement" gap
CLUSTER_MIN_KEYWORDS = 3        # competitor must win ≥3 kw in offering to be a cluster

# Severity scoring weights (1-10 scale).
SEV_BASE_COVERAGE     = 3
SEV_BASE_DISPLACEMENT = 4   # we have a page, so easier to fix → higher prio
SEV_BONUS_PER_TOP10_COMP = 0.4   # each tracked competitor in top 10 adds
SEV_GSC_IMPR_BONUS      = 2      # ≥100 impressions in 14d
SEV_GSC_CLICKS_BONUS    = 3      # any clicks in 14d (real traffic at stake)
SEV_PRIMARY_THREAT_BONUS = 1     # primary-threat competitor in top 10

OUTPUTS_REPORTS = Path(__file__).resolve().parent.parent / "outputs" / "reports"
OUTPUTS_AUDITS  = Path(__file__).resolve().parent.parent / "outputs" / "audits"


# ---------------------------------------------------------------------------
# Read phase
# ---------------------------------------------------------------------------

def load_offerings() -> list[str]:
    rows = fetch_all(
        "SELECT DISTINCT offering FROM keywords WHERE status='active' AND offering IS NOT NULL "
        "ORDER BY offering"
    )
    return [r["offering"] for r in rows]


def load_offering_data(offering: str) -> dict:
    """
    Pulls everything needed for gap analysis on one offering:
      - keywords + their latest Damco position (DataForSEO) + GSC stats
      - top-10 competitors per keyword (latest snapshot)
    """
    # Damco + GSC view, one row per keyword
    damco_rows = fetch_all(
        """
        WITH latest_serp AS (
            SELECT keyword_id, max(date) AS d FROM keyword_rankings
             WHERE source = 'dataforseo' GROUP BY keyword_id
        ),
        latest_gsc AS (
            SELECT keyword_id, max(date) AS d FROM keyword_rankings
             WHERE source = 'gsc' GROUP BY keyword_id
        )
        SELECT k.id, k.keyword, k.target_url,
               serp.rank_position AS damco_position,
               serp.url_found     AS damco_url,
               gsc.rank_position  AS gsc_position,
               gsc.clicks         AS gsc_clicks,
               gsc.impressions    AS gsc_impressions,
               gsc.ctr            AS gsc_ctr
          FROM keywords k
     LEFT JOIN latest_serp ls ON ls.keyword_id = k.id
     LEFT JOIN keyword_rankings serp ON serp.keyword_id = k.id AND serp.date = ls.d AND serp.source = 'dataforseo'
     LEFT JOIN latest_gsc lg ON lg.keyword_id = k.id
     LEFT JOIN keyword_rankings gsc ON gsc.keyword_id = k.id AND gsc.date = lg.d AND gsc.source = 'gsc'
         WHERE k.status = 'active' AND k.offering = %s
         ORDER BY k.keyword
        """,
        [offering],
    )

    # Top-10 competitor rows, scoped to keywords in this offering
    competitor_rows = fetch_all(
        """
        WITH latest AS (
            SELECT keyword_id, max(date) AS d FROM competitor_rankings GROUP BY keyword_id
        )
        SELECT cr.keyword_id, cr.rank_position, cr.url_found, cr.url_title,
               cr.page_type, c.competitor_domain, c.company_name,
               c.category, c.threat_tier
          FROM competitor_rankings cr
          JOIN latest l ON l.keyword_id = cr.keyword_id AND l.d = cr.date
          JOIN keywords k ON k.id = cr.keyword_id
          JOIN competitors c ON c.id = cr.competitor_id
         WHERE k.status = 'active' AND k.offering = %s
           AND cr.rank_position BETWEEN 1 AND 10
           AND c.is_tracked = TRUE
         ORDER BY cr.keyword_id, cr.rank_position
        """,
        [offering],
    )

    # Group competitors by keyword_id
    by_kw: dict[int, list[dict]] = defaultdict(list)
    for r in competitor_rows:
        by_kw[r["keyword_id"]].append(r)

    return {
        "offering":      offering,
        "keywords":      damco_rows,
        "competitors_by_keyword": dict(by_kw),
    }


# ---------------------------------------------------------------------------
# Detection
# ---------------------------------------------------------------------------

def classify(damco_pos: int | None, competitors: list[dict]) -> tuple[str, int]:
    """
    Returns (gap_type, base_severity) for one keyword.
        - 'none'         — Damco in top 10
        - 'coverage_gap' — Damco missing from top 100
        - 'displacement' — Damco 11-30
        - 'low_priority' — Damco 31+ (lower-prio displacement)
    """
    if damco_pos is not None and 1 <= damco_pos <= 10:
        return "none", 0
    if damco_pos is None:
        if competitors:
            return "coverage_gap", SEV_BASE_COVERAGE
        return "none", 0  # nobody ranks here, weird keyword
    if DISPLACEMENT_RANGE[0] <= damco_pos <= DISPLACEMENT_RANGE[1]:
        return "displacement", SEV_BASE_DISPLACEMENT
    return "low_priority", 1


def compute_severity(base: int, kw_row: dict, competitors: list[dict]) -> float:
    """Layer GSC-traffic + competitor-count bonuses on the base severity."""
    if base == 0:
        return 0
    score = float(base)
    score += min(len(competitors), 10) * SEV_BONUS_PER_TOP10_COMP

    impr = (kw_row.get("gsc_impressions") or 0)
    clicks = (kw_row.get("gsc_clicks") or 0)
    if impr >= 100:
        score += SEV_GSC_IMPR_BONUS
    if clicks > 0:
        score += SEV_GSC_CLICKS_BONUS

    if any(c.get("threat_tier") == "primary" for c in competitors):
        score += SEV_PRIMARY_THREAT_BONUS

    return round(score, 1)


def detect_cluster_wins(data: dict) -> list[dict]:
    """Competitors winning ≥CLUSTER_MIN_KEYWORDS top-10 placements in this offering."""
    win_counter: Counter[str] = Counter()
    win_keywords: dict[str, list[str]] = defaultdict(list)
    win_positions: dict[str, list[int]] = defaultdict(list)

    kw_by_id = {kw["id"]: kw["keyword"] for kw in data["keywords"]}
    for kid, comps in data["competitors_by_keyword"].items():
        kw_text = kw_by_id.get(kid)
        if not kw_text:
            continue
        for c in comps:
            d = c["competitor_domain"]
            win_counter[d] += 1
            win_keywords[d].append(kw_text)
            win_positions[d].append(c["rank_position"])

    clusters = []
    for domain, count in win_counter.most_common():
        if count < CLUSTER_MIN_KEYWORDS:
            break
        positions = win_positions[domain]
        clusters.append({
            "competitor_domain": domain,
            "wins":              count,
            "avg_position":      round(sum(positions) / len(positions), 2),
            "best_position":     min(positions),
            "sample_keywords":   win_keywords[domain][:8],
        })
    return clusters


def build_gap_records(data: dict) -> list[dict]:
    """One record per keyword — full classification + scoring."""
    out: list[dict] = []
    for kw in data["keywords"]:
        comps = data["competitors_by_keyword"].get(kw["id"], [])
        gap_type, base = classify(kw["damco_position"], comps)
        severity = compute_severity(base, kw, comps)
        # Pick top 3 competitors for display
        top3 = comps[:3]
        out.append({
            "keyword":           kw["keyword"],
            "target_url":        kw["target_url"],
            "damco_position":    kw["damco_position"],
            "damco_url":         kw["damco_url"],
            "gsc_position":      kw["gsc_position"],
            "gsc_clicks":        kw["gsc_clicks"] or 0,
            "gsc_impressions":   kw["gsc_impressions"] or 0,
            "gap_type":          gap_type,
            "severity":          severity,
            "top10_comp_count":  len(comps),
            "top_competitors":   [
                {"pos": c["rank_position"], "domain": c["competitor_domain"],
                 "category": c["category"], "threat_tier": c["threat_tier"]}
                for c in top3
            ],
        })
    return out


# ---------------------------------------------------------------------------
# LLM narrative (optional)
# ---------------------------------------------------------------------------

def make_narrative_prompt(offering: str, gaps: list[dict],
                          clusters: list[dict], totals: dict) -> str:
    """Prompt template for the per-offering executive summary + recs."""
    top_coverage   = [g for g in gaps if g["gap_type"] == "coverage_gap"]
    top_coverage.sort(key=lambda g: -g["severity"])
    top_displaced  = [g for g in gaps if g["gap_type"] == "displacement"]
    top_displaced.sort(key=lambda g: -g["severity"])

    def fmt_gap(g):
        comps = ", ".join(f"{c['domain']} (#{c['pos']})" for c in g["top_competitors"][:3])
        return (f"  - {g['keyword']!r}: damco pos={g['damco_position']}, "
                f"gsc_clicks={g['gsc_clicks']}, gsc_impr={g['gsc_impressions']}, "
                f"top competitors: {comps}")

    lines = [
        f"Offering: {offering}",
        f"Total active keywords in offering: {totals['total']}",
        f"Damco in top 10:      {totals['in_top10']}",
        f"Coverage gaps:        {totals['coverage_gap']}",
        f"Displacement gaps:    {totals['displacement']}",
        f"Low-priority (rank >30): {totals['low_priority']}",
        "",
        f"Top 15 COVERAGE GAPS (Damco not in top 100, competitors in top 10):",
    ]
    lines.extend(fmt_gap(g) for g in top_coverage[:15])
    lines.append("")
    lines.append(f"Top 10 DISPLACEMENT gaps (Damco close, competitor in top 10):")
    lines.extend(fmt_gap(g) for g in top_displaced[:10])
    lines.append("")
    lines.append(f"CLUSTER WINS (competitors dominating this offering):")
    for c in clusters[:6]:
        lines.append(
            f"  - {c['competitor_domain']}: wins {c['wins']} keywords, "
            f"avg pos {c['avg_position']}, best #{c['best_position']}. "
            f"Sample: {', '.join(c['sample_keywords'][:5])}"
        )

    body = "\n".join(lines)

    return f"""You are an SEO strategist briefing Damco's marketing team. Below is competitive gap data for one offering. Generate:

1. A 2-3 paragraph executive summary of the competitive landscape and where Damco stands.
2. Top 5 prioritized recommendations as a numbered list. For each: state the action, name the specific keywords/URLs/competitors involved, and label it as either "QUICK WIN" (achievable in 2-4 weeks via on-page work) or "INVESTMENT" (requires new content/pages, 1-3 months).
3. A one-line "headline takeaway" suitable for a stakeholder email.

Be concrete. Reference specific keywords and competitor domains from the data. Don't hedge. If something looks dominated and unwinnable, say so.

GAP DATA:
{body}

End your response. No additional commentary."""


def get_narrative(offering: str, gaps: list[dict], clusters: list[dict],
                  totals: dict) -> tuple[str | None, dict | None]:
    """Returns (narrative_text, usage) or (None, None) when LLM unavailable."""
    prompt = make_narrative_prompt(offering, gaps, clusters, totals)
    try:
        text, usage = call_claude(prompt, tier="default", max_tokens=2500)
        return text, usage
    except LLMUnavailableError as exc:
        logger.warning("LLM narrative unavailable for %s: %s", offering, exc)
        return None, None


def rule_based_narrative(offering: str, gaps: list[dict],
                         clusters: list[dict], totals: dict) -> str:
    """Plain rule-based fallback when LLM isn't available."""
    coverage = sorted([g for g in gaps if g["gap_type"] == "coverage_gap"],
                      key=lambda g: -g["severity"])[:10]
    displace = sorted([g for g in gaps if g["gap_type"] == "displacement"],
                      key=lambda g: -g["severity"])[:10]

    lines = [
        f"## Executive summary (rule-based — load Anthropic credit for narrative)",
        "",
        f"- **{totals['total']}** active keywords in `{offering}`.",
        f"- **{totals['in_top10']}** ({totals['in_top10']*100//max(1,totals['total'])}%) on page 1 — Damco's defended territory.",
        f"- **{totals['coverage_gap']}** coverage gaps (competitor in top 10, Damco missing entirely).",
        f"- **{totals['displacement']}** displacement gaps (Damco at #11-30, competitor on page 1).",
        f"- **{len(clusters)}** competitor cluster(s) dominating ≥{CLUSTER_MIN_KEYWORDS} keywords each.",
        "",
        "## Top quick-win candidates (displacement, by severity)",
        "",
    ]
    if not displace:
        lines.append("_None._")
    for g in displace:
        comps = ", ".join(f"{c['domain']} (#{c['pos']})" for c in g["top_competitors"][:3])
        lines.append(f"- `{g['keyword']}` — Damco #{g['damco_position']}, "
                     f"{g['gsc_clicks']} clicks / {g['gsc_impressions']} impr. Top: {comps}")

    lines.extend(["", "## Top content-investment candidates (coverage gap, by severity)", ""])
    if not coverage:
        lines.append("_None._")
    for g in coverage:
        comps = ", ".join(f"{c['domain']} (#{c['pos']})" for c in g["top_competitors"][:3])
        target = g["target_url"] or "(no target_url assigned)"
        lines.append(f"- `{g['keyword']}` — {g['gsc_clicks']} clicks / {g['gsc_impressions']} impr. "
                     f"Suggested target: `{target}`. Owns top 10: {comps}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def write_markdown(offering: str, gaps: list[dict], clusters: list[dict],
                   totals: dict, narrative: str | None,
                   narrative_usage: dict | None) -> Path:
    OUTPUTS_AUDITS.mkdir(parents=True, exist_ok=True)
    safe_offering = offering.replace(" ", "_").replace("/", "-").replace(",", "")
    path = OUTPUTS_AUDITS / f"gap_analysis_{safe_offering}_{date.today().isoformat()}.md"

    parts: list[str] = []
    parts.append(f"# Competitive Gap Analysis — `{offering}`")
    parts.append("")
    parts.append(f"_Generated {date.today().isoformat()} by `{AGENT_NAME}`._")
    parts.append("")
    if narrative_usage:
        parts.append(f"_LLM narrative: {narrative_usage['model']}, "
                     f"in={narrative_usage['input_tokens']}, "
                     f"out={narrative_usage['output_tokens']}, "
                     f"~${narrative_usage['est_cost_usd']:.4f}_")
        parts.append("")
    parts.append("---")
    parts.append("")
    if narrative:
        parts.append("## Strategic narrative")
        parts.append("")
        parts.append(narrative)
        parts.append("")
        parts.append("---")
        parts.append("")
    else:
        parts.append(rule_based_narrative(offering, gaps, clusters, totals))
        parts.append("")
        parts.append("---")
        parts.append("")

    # Rule-based detail tables follow regardless of LLM availability
    parts.append("## Gap counts")
    parts.append("")
    parts.append(f"| Type | Count |")
    parts.append(f"|---|---:|")
    parts.append(f"| Active keywords | {totals['total']} |")
    parts.append(f"| Damco in top 10 (no gap) | {totals['in_top10']} |")
    parts.append(f"| Coverage gap (Damco not in top 100) | {totals['coverage_gap']} |")
    parts.append(f"| Displacement (Damco #11-30) | {totals['displacement']} |")
    parts.append(f"| Low priority (Damco rank > 30) | {totals['low_priority']} |")
    parts.append("")

    if clusters:
        parts.append("## Competitor cluster wins")
        parts.append("")
        parts.append("Competitors that win top-10 placements for ≥3 keywords in this offering.")
        parts.append("")
        parts.append("| Competitor | Wins | Avg pos | Best pos | Sample keywords |")
        parts.append("|---|---:|---:|---:|---|")
        for c in clusters:
            samp = ", ".join(f"`{k}`" for k in c["sample_keywords"][:5])
            parts.append(f"| `{c['competitor_domain']}` | {c['wins']} | {c['avg_position']} | #{c['best_position']} | {samp} |")
        parts.append("")

    # Full lists
    coverage = sorted([g for g in gaps if g["gap_type"] == "coverage_gap"],
                      key=lambda g: -g["severity"])
    if coverage:
        parts.append(f"## All coverage gaps ({len(coverage)})")
        parts.append("")
        parts.append("Damco doesn't rank in the top 100 but at least one tracked competitor sits in the top 10.")
        parts.append("")
        parts.append("| Severity | Keyword | GSC clicks | GSC impr | Top competitors |")
        parts.append("|---:|---|---:|---:|---|")
        for g in coverage[:50]:
            comps = ", ".join(f"`{c['domain']}` (#{c['pos']})" for c in g["top_competitors"])
            parts.append(f"| {g['severity']} | `{g['keyword']}` | {g['gsc_clicks']} | {g['gsc_impressions']} | {comps} |")
        if len(coverage) > 50:
            parts.append(f"\n_...{len(coverage) - 50} more in the Excel report._")
        parts.append("")

    displaced = sorted([g for g in gaps if g["gap_type"] == "displacement"],
                       key=lambda g: -g["severity"])
    if displaced:
        parts.append(f"## All displacement gaps ({len(displaced)})")
        parts.append("")
        parts.append("Damco ranks #11-30 — page exists, optimization push could land top 10.")
        parts.append("")
        parts.append("| Severity | Keyword | Damco pos | Damco URL | GSC clicks | Top competitors |")
        parts.append("|---:|---|---:|---|---:|---|")
        for g in displaced[:50]:
            comps = ", ".join(f"`{c['domain']}` (#{c['pos']})" for c in g["top_competitors"])
            damco_url = (g["damco_url"] or "—")
            if len(damco_url) > 45:
                damco_url = damco_url[:42] + "..."
            parts.append(f"| {g['severity']} | `{g['keyword']}` | #{g['damco_position']} | `{damco_url}` | {g['gsc_clicks']} | {comps} |")
        parts.append("")

    path.write_text("\n".join(parts), encoding="utf-8")
    return path


def write_excel(all_offering_data: list[tuple[str, list[dict], list[dict], dict]]) -> Path:
    """One workbook covering all offerings analyzed in this run."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    OUTPUTS_REPORTS.mkdir(parents=True, exist_ok=True)
    path = OUTPUTS_REPORTS / f"gap_analysis_{date.today().isoformat()}.xlsx"

    wb = openpyxl.Workbook()
    # ---- Sheet 1: per-keyword classification ----
    ws = wb.active
    ws.title = "Per-Keyword"
    headers = [
        "Offering", "Keyword", "Gap Type", "Severity",
        "Damco Position", "Damco URL", "Target URL",
        "GSC Position", "GSC Clicks (14d)", "GSC Impressions (14d)",
        "Top-10 Competitor Count",
        "Comp #1 Pos", "Comp #1 Domain",
        "Comp #2 Pos", "Comp #2 Domain",
        "Comp #3 Pos", "Comp #3 Domain",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill

    for offering, gaps, _clusters, _totals in all_offering_data:
        for g in gaps:
            comps = g["top_competitors"]
            row = [
                offering, g["keyword"], g["gap_type"], g["severity"],
                g["damco_position"] or "NF",
                g["damco_url"] or "",
                g["target_url"] or "",
                g["gsc_position"] or "",
                g["gsc_clicks"] or 0,
                g["gsc_impressions"] or 0,
                g["top10_comp_count"],
            ]
            for i in range(3):
                if i < len(comps):
                    row.append(comps[i]["pos"])
                    row.append(comps[i]["domain"])
                else:
                    row.extend(["", ""])
            ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_letter, width in zip(
        ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q"],
        [38, 50, 14, 9, 10, 50, 50, 10, 11, 13, 8, 8, 26, 8, 26, 8, 26],
    ):
        ws.column_dimensions[col_letter].width = width

    # ---- Sheet 2: cluster wins ----
    ws2 = wb.create_sheet("Cluster Wins")
    ws2.append(["Offering", "Competitor Domain", "Wins", "Avg Position", "Best Position", "Sample Keywords"])
    for col in range(1, 7):
        c = ws2.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
    for offering, _gaps, clusters, _totals in all_offering_data:
        for c in clusters:
            ws2.append([
                offering, c["competitor_domain"], c["wins"],
                c["avg_position"], c["best_position"],
                ", ".join(c["sample_keywords"][:8]),
            ])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    for col_letter, width in zip("ABCDEF", [38, 38, 7, 11, 13, 95]):
        ws2.column_dimensions[col_letter].width = width

    # ---- Sheet 3: offering summary ----
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Offering", "Total Keywords", "In Top 10", "Coverage Gaps",
                "Displacement Gaps", "Low Priority", "Cluster-Win Competitors"])
    for col in range(1, 8):
        c = ws3.cell(row=1, column=col)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
    for offering, _gaps, clusters, totals in all_offering_data:
        ws3.append([
            offering, totals["total"], totals["in_top10"],
            totals["coverage_gap"], totals["displacement"], totals["low_priority"],
            len(clusters),
        ])
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions
    for col_letter, width in zip("ABCDEFG", [38, 14, 11, 14, 18, 13, 22]):
        ws3.column_dimensions[col_letter].width = width

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Per-offering pipeline
# ---------------------------------------------------------------------------

def analyze_offering(offering: str, with_narrative: bool) -> tuple[
    list[dict], list[dict], dict, Path, dict | None
]:
    """Returns (gaps, clusters, totals, markdown_path, narrative_usage)."""
    data = load_offering_data(offering)
    gaps = build_gap_records(data)
    clusters = detect_cluster_wins(data)

    counter = Counter(g["gap_type"] for g in gaps)
    totals = {
        "total":          len(gaps),
        "in_top10":       counter.get("none", 0),
        "coverage_gap":   counter.get("coverage_gap", 0),
        "displacement":   counter.get("displacement", 0),
        "low_priority":   counter.get("low_priority", 0),
    }
    in_top10_with_pos = sum(1 for g in gaps if g["damco_position"] and g["damco_position"] <= 10)
    totals["in_top10"] = in_top10_with_pos
    # Recount: coverage = damco_position is None AND >=1 competitor in top 10
    # The Counter approach already covers this, so leave totals as is for non-top10 categories.

    narrative: str | None = None
    narrative_usage: dict | None = None
    if with_narrative:
        logger.info("Generating LLM narrative for %s...", offering)
        narrative, narrative_usage = get_narrative(offering, gaps, clusters, totals)

    md_path = write_markdown(offering, gaps, clusters, totals, narrative, narrative_usage)
    return gaps, clusters, totals, md_path, narrative_usage


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(offering: str | None = None, with_narrative: bool = False,
        dry_run: bool = False) -> dict:
    start = time.monotonic()
    target_offerings = [offering] if offering else load_offerings()
    logger.info("Analyzing %d offering(s): %s%s",
                len(target_offerings),
                ", ".join(target_offerings[:5]) + ("..." if len(target_offerings) > 5 else ""),
                "  [LLM narrative ON]" if with_narrative else "")

    all_offering_data: list[tuple[str, list[dict], list[dict], dict]] = []
    total_llm_cost = 0.0
    total_llm_in = total_llm_out = 0
    llm_offerings_done = 0
    md_paths: list[Path] = []

    for offering_name in target_offerings:
        logger.info("--- %s ---", offering_name)
        gaps, clusters, totals, md_path, narrative_usage = analyze_offering(
            offering_name, with_narrative,
        )
        all_offering_data.append((offering_name, gaps, clusters, totals))
        md_paths.append(md_path)
        if narrative_usage:
            total_llm_cost += narrative_usage["est_cost_usd"]
            total_llm_in   += narrative_usage["input_tokens"]
            total_llm_out  += narrative_usage["output_tokens"]
            llm_offerings_done += 1
        logger.info("%s: %d total, %d coverage gaps, %d displacement, %d cluster wins",
                    offering_name, totals["total"], totals["coverage_gap"],
                    totals["displacement"], len(clusters))

    xlsx_path = write_excel(all_offering_data)
    duration = time.monotonic() - start

    # Aggregate counters
    total_kw         = sum(t[3]["total"]        for t in all_offering_data)
    total_coverage   = sum(t[3]["coverage_gap"] for t in all_offering_data)
    total_displaced  = sum(t[3]["displacement"] for t in all_offering_data)
    total_clusters   = sum(len(t[2])            for t in all_offering_data)

    if not dry_run:
        record_agent_run(
            agent_name=AGENT_NAME,
            status="success",
            records_processed=total_kw,
            errors=[],
            duration_seconds=round(duration, 2),
            metadata={
                "run_date":            date.today().isoformat(),
                "offerings":           target_offerings,
                "offering_count":      len(target_offerings),
                "with_narrative":      with_narrative,
                "llm_offerings_done":  llm_offerings_done,
                "llm_input_tokens":    total_llm_in,
                "llm_output_tokens":   total_llm_out,
                "llm_est_cost_usd":    round(total_llm_cost, 4),
                "total_keywords":      total_kw,
                "total_coverage_gaps": total_coverage,
                "total_displacement":  total_displaced,
                "total_clusters":      total_clusters,
                "xlsx_path":           str(xlsx_path),
                "markdown_paths":      [str(p) for p in md_paths],
            },
        )

    # Console summary
    print()
    print(f"  {'=' * 72}")
    print(f"   COMPETITIVE GAP ANALYZER — {date.today().isoformat()}")
    print(f"  {'=' * 72}")
    print()
    print(f"  Offerings analyzed:    {len(target_offerings)}")
    print(f"  Active keywords:       {total_kw}")
    print(f"  Coverage gaps:         {total_coverage}")
    print(f"  Displacement gaps:     {total_displaced}")
    print(f"  Competitor clusters:   {total_clusters}")
    if with_narrative:
        print(f"  LLM offerings done:    {llm_offerings_done}/{len(target_offerings)}")
        print(f"  LLM total cost:        ~${total_llm_cost:.4f}")
    print(f"  Excel report:          {xlsx_path}")
    print(f"  Markdown reports:      {len(md_paths)} file(s) under outputs/audits/")
    print(f"  Duration:              {duration:.1f}s")
    print()

    return {
        "status":           "success",
        "offering_count":   len(target_offerings),
        "total_keywords":   total_kw,
        "coverage_gaps":    total_coverage,
        "displacement":     total_displaced,
        "clusters":         total_clusters,
        "xlsx_path":        str(xlsx_path),
        "markdown_paths":   [str(p) for p in md_paths],
        "llm_cost_usd":     round(total_llm_cost, 4),
        "duration_seconds": round(duration, 2),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Damco Competitive Gap Analyzer")
    parser.add_argument("--offering", help="Restrict to one offering (default: all)")
    parser.add_argument("--with-narrative", action="store_true",
                        help="Generate LLM-narrated executive summary + recommendations. "
                             "Requires ANTHROPIC_API_KEY + credit. Falls back to "
                             "rule-based summaries when LLM is unavailable.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Generate reports but skip agent_runs DB write")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(name)s  %(levelname)s  %(message)s",
    )

    run(offering=args.offering, with_narrative=args.with_narrative, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
