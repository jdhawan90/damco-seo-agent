"""
Backlink Analyzer — Phase 2 module of Competitive Intelligence
===============================================================

Profiles each tracked competitor's backlink portfolio via the DataForSEO
Backlinks API. Surfaces:

  - Top referring domains per primary threat
  - Sites that link to MULTIPLE primary threats (highest-leverage outreach
    targets — they already publish about this space, just don't link to us)
  - Anchor-text patterns competitors are building
  - Domain-authority distribution per competitor

State is stored in `competitor_backlinks` (migration 008). The DataForSEO
Backlinks API needs a separate subscription from the SERP pay-per-query
endpoint; the module degrades gracefully (no crash, clear error message)
when access is denied.

Usage
-----
    # Default: primary threat competitors, top 500 backlinks each, monthly cadence
    python -m competitive_intelligence.backlink_analyzer

    # Also include watch-tier competitors (wider scope)
    python -m competitive_intelligence.backlink_analyzer --threat-tier primary,watch

    # One specific competitor
    python -m competitive_intelligence.backlink_analyzer --domain itransition.com

    # Lower limit to control cost
    python -m competitive_intelligence.backlink_analyzer --limit 100

    # Force re-pull ignoring 30-day cadence
    python -m competitive_intelligence.backlink_analyzer --all

    # Dry-run: still calls the API (real cost), but doesn't write to DB
    python -m competitive_intelligence.backlink_analyzer --dry-run

    # Generate the analysis report from already-stored data, no API calls
    python -m competitive_intelligence.backlink_analyzer --analyze-only
"""

from __future__ import annotations

import argparse
import json
import logging
import sys
import time
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from common.database import fetch_all, record_agent_run, connection
from common.connectors.dataforseo import (
    get_backlinks, DataForSEOAccessDenied, DataForSEOError,
)


logger = logging.getLogger("backlink_analyzer")
AGENT_NAME = "competitive_intelligence.backlink_analyzer"

DEFAULT_THREAT_TIERS = ("primary",)
DEFAULT_LIMIT = 500
DEFAULT_CADENCE_DAYS = 30   # backlink data doesn't change rapidly

OUTPUT_AUDITS  = Path(__file__).resolve().parent.parent / "outputs" / "audits"
OUTPUT_REPORTS = Path(__file__).resolve().parent.parent / "outputs" / "reports"


# ---------------------------------------------------------------------------
# Read phase
# ---------------------------------------------------------------------------

def load_target_competitors(threat_tiers: tuple[str, ...],
                            domain: str | None,
                            cadence_days: int,
                            force_all: bool) -> list[dict]:
    """Pick which competitors to pull backlinks for."""
    if domain:
        rows = fetch_all(
            "SELECT id, competitor_domain, company_name, category, threat_tier, "
            "       (SELECT max(date_discovered) FROM competitor_backlinks "
            "         WHERE competitor_id = c.id) AS last_pull "
            "  FROM competitors c WHERE c.competitor_domain = %s AND c.is_tracked = TRUE",
            [domain],
        )
        if not rows:
            logger.warning("No tracked competitor with domain=%s", domain)
            return []
        return rows

    rows = fetch_all(
        """
        SELECT c.id, c.competitor_domain, c.company_name, c.category, c.threat_tier,
               (SELECT max(date_discovered) FROM competitor_backlinks
                 WHERE competitor_id = c.id) AS last_pull
          FROM competitors c
         WHERE c.threat_tier = ANY(%s) AND c.is_tracked = TRUE
         ORDER BY c.keyword_appearance_count DESC NULLS LAST, c.competitor_domain
        """,
        [list(threat_tiers)],
    )

    if force_all:
        return rows

    today = date.today()
    due: list[dict] = []
    for r in rows:
        if r["last_pull"] is None or (today - r["last_pull"]).days >= cadence_days:
            due.append(r)
    return due


# ---------------------------------------------------------------------------
# Fetch + store
# ---------------------------------------------------------------------------

def fetch_and_store(competitor: dict, limit: int, dry_run: bool) -> dict:
    """Pull backlinks for one competitor; upsert into competitor_backlinks."""
    domain = competitor["competitor_domain"]
    counters = {"fetched": 0, "inserted": 0, "skipped_existing": 0, "error": None}

    try:
        items = get_backlinks(domain, limit=limit, mode="as_is")
    except DataForSEOAccessDenied as exc:
        counters["error"] = "access_denied"
        counters["error_detail"] = str(exc)
        return counters
    except DataForSEOError as exc:
        counters["error"] = "api_error"
        counters["error_detail"] = str(exc)
        return counters

    counters["fetched"] = len(items)
    if dry_run or not items:
        return counters

    with connection() as conn:
        with conn.cursor() as cur:
            for it in items:
                cur.execute(
                    """
                    INSERT INTO competitor_backlinks
                        (competitor_id, source_url, source_domain, target_url,
                         anchor_text, is_dofollow, domain_rank,
                         first_seen, last_seen, raw_payload)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
                    ON CONFLICT (competitor_id, source_url, data_source) DO UPDATE
                       SET last_seen   = EXCLUDED.last_seen,
                           domain_rank = EXCLUDED.domain_rank,
                           anchor_text = EXCLUDED.anchor_text,
                           raw_payload = EXCLUDED.raw_payload
                    """,
                    (
                        competitor["id"],
                        it.get("source_url"),
                        it.get("source_domain"),
                        it.get("target_url"),
                        it.get("anchor"),
                        bool(it.get("dofollow")) if it.get("dofollow") is not None else None,
                        it.get("rank"),
                        _safe_date(it.get("first_seen")),
                        _safe_date(it.get("last_seen")),
                        json.dumps(it.get("raw") or {}),
                    ),
                )
                if cur.rowcount > 0:
                    counters["inserted"] += 1
        conn.commit()
    return counters


def _safe_date(s: str | None) -> date | None:
    """DataForSEO returns dates as ISO strings or ISO datetimes."""
    if not s:
        return None
    try:
        # Strip time portion if present
        return date.fromisoformat(s[:10])
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Analysis
# ---------------------------------------------------------------------------

def compute_analysis(competitor_ids: list[int]) -> dict:
    """
    Cross-analyzes backlinks across the given competitors. Returns:
      - per_competitor_summary
      - top_referring_domains_overall
      - intersection (referring domains linking to ≥2 of the competitors)
      - anchor_distribution
    """
    if not competitor_ids:
        return {
            "per_competitor": [],
            "top_referring_domains_overall": [],
            "intersection": [],
            "anchor_distribution": [],
        }

    per_comp = fetch_all(
        """
        SELECT c.id, c.competitor_domain, c.threat_tier, c.category,
               count(*) AS total_links,
               count(DISTINCT cb.source_domain) AS unique_referring_domains,
               count(*) FILTER (WHERE cb.is_dofollow) AS dofollow_links,
               round(avg(cb.domain_rank)::numeric, 1) AS avg_rank,
               max(cb.domain_rank) AS max_rank
          FROM competitor_backlinks cb
          JOIN competitors c ON c.id = cb.competitor_id
         WHERE c.id = ANY(%s)
         GROUP BY c.id, c.competitor_domain, c.threat_tier, c.category
         ORDER BY total_links DESC
        """,
        [competitor_ids],
    )

    top_referrers = fetch_all(
        """
        SELECT source_domain,
               count(DISTINCT competitor_id) AS competitors_linked,
               count(*) AS total_links,
               max(domain_rank) AS max_rank,
               array_agg(DISTINCT (SELECT competitor_domain FROM competitors WHERE id = cb.competitor_id)) AS to_domains
          FROM competitor_backlinks cb
         WHERE competitor_id = ANY(%s)
         GROUP BY source_domain
         ORDER BY competitors_linked DESC, total_links DESC
         LIMIT 50
        """,
        [competitor_ids],
    )

    # "Intersection" = referring domains linking to ≥2 of our primary threats.
    # These are the operational gold for outreach: they already publish about
    # this space.
    intersection = [r for r in top_referrers if r["competitors_linked"] >= 2]

    anchors = fetch_all(
        """
        SELECT lower(trim(anchor_text)) AS anchor, count(*) AS n,
               count(DISTINCT competitor_id) AS competitors_using
          FROM competitor_backlinks
         WHERE competitor_id = ANY(%s) AND anchor_text IS NOT NULL AND length(trim(anchor_text)) > 0
         GROUP BY lower(trim(anchor_text))
         ORDER BY n DESC LIMIT 30
        """,
        [competitor_ids],
    )

    return {
        "per_competitor":              per_comp,
        "top_referring_domains_overall": top_referrers,
        "intersection":                intersection,
        "anchor_distribution":         anchors,
    }


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def write_excel(analysis: dict, competitors: list[dict]) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    OUTPUT_REPORTS.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_REPORTS / f"backlink_analysis_{date.today().isoformat()}.xlsx"
    wb = openpyxl.Workbook()
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    def header_row(ws, headers):
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = header_font
            c.fill = header_fill

    # Sheet 1: per-competitor summary
    ws = wb.active
    ws.title = "Per-Competitor"
    header_row(ws, ["Domain", "Tier", "Category", "Total Links",
                    "Unique Referring Domains", "Dofollow", "Avg Rank", "Max Rank"])
    for r in analysis["per_competitor"]:
        ws.append([
            r["competitor_domain"], r["threat_tier"], r["category"] or "",
            r["total_links"], r["unique_referring_domains"], r["dofollow_links"],
            float(r["avg_rank"]) if r["avg_rank"] else 0,
            r["max_rank"] or 0,
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col, w in zip("ABCDEFGH", [32, 12, 14, 12, 24, 12, 10, 10]):
        ws.column_dimensions[col].width = w

    # Sheet 2: top referring domains overall (with intersection flag)
    ws2 = wb.create_sheet("Top Referring Domains")
    header_row(ws2, ["Source Domain", "# Primary Threats Linking", "Total Links",
                     "Max Rank", "Linked Competitors"])
    for r in analysis["top_referring_domains_overall"]:
        ws2.append([
            r["source_domain"], r["competitors_linked"], r["total_links"],
            r["max_rank"] or 0,
            ", ".join(sorted(r["to_domains"] or [])),
        ])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    for col, w in zip("ABCDE", [40, 24, 12, 10, 70]):
        ws2.column_dimensions[col].width = w

    # Sheet 3: outreach prospects (referring domains linking to ≥2 competitors)
    ws3 = wb.create_sheet("Outreach Prospects")
    header_row(ws3, ["Source Domain", "# Competitors Linking", "Total Links",
                     "Max Rank", "Linked Competitors"])
    for r in analysis["intersection"]:
        ws3.append([
            r["source_domain"], r["competitors_linked"], r["total_links"],
            r["max_rank"] or 0,
            ", ".join(sorted(r["to_domains"] or [])),
        ])
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = ws3.dimensions
    for col, w in zip("ABCDE", [40, 22, 12, 10, 70]):
        ws3.column_dimensions[col].width = w

    # Sheet 4: anchor text distribution
    ws4 = wb.create_sheet("Anchor Patterns")
    header_row(ws4, ["Anchor Text", "Total Uses", "# Competitors Using"])
    for r in analysis["anchor_distribution"]:
        ws4.append([r["anchor"], r["n"], r["competitors_using"]])
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = ws4.dimensions
    for col, w in zip("ABC", [60, 12, 22]):
        ws4.column_dimensions[col].width = w

    wb.save(path)
    return path


def write_markdown(analysis: dict, fetch_summary: dict) -> Path:
    OUTPUT_AUDITS.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_AUDITS / f"backlink_analysis_{date.today().isoformat()}.md"

    parts: list[str] = []
    parts.append(f"# Competitor Backlink Analysis — {date.today().isoformat()}")
    parts.append("")
    parts.append(f"_Generated by `{AGENT_NAME}`._")
    parts.append("")

    parts.append("## Fetch summary")
    parts.append("")
    parts.append(f"| Metric | Value |")
    parts.append(f"|---|---:|")
    parts.append(f"| Competitors targeted | {fetch_summary['competitors_targeted']} |")
    parts.append(f"| Successful pulls | {fetch_summary['successful_pulls']} |")
    parts.append(f"| Access denied / errors | {fetch_summary['failed_pulls']} |")
    parts.append(f"| Total backlinks fetched | {fetch_summary['total_fetched']} |")
    parts.append(f"| Total backlinks inserted | {fetch_summary['total_inserted']} |")
    parts.append("")

    if not analysis["per_competitor"]:
        parts.append("_No backlink data to analyze. Either no API access, or no competitors with stored data yet._")
        path.write_text("\n".join(parts), encoding="utf-8")
        return path

    parts.append("## Per-competitor backlink profile")
    parts.append("")
    parts.append("| Competitor | Tier | Category | Total Links | Unique Referring Domains | Dofollow | Avg Rank | Max Rank |")
    parts.append("|---|---|---|---:|---:|---:|---:|---:|")
    for r in analysis["per_competitor"]:
        parts.append(f"| `{r['competitor_domain']}` | {r['threat_tier']} | {r['category'] or '?'} | "
                     f"{r['total_links']} | {r['unique_referring_domains']} | {r['dofollow_links']} | "
                     f"{r['avg_rank'] or 0} | {r['max_rank'] or 0} |")
    parts.append("")

    parts.append(f"## Outreach prospects — domains linking to ≥2 primary threats ({len(analysis['intersection'])})")
    parts.append("")
    parts.append("These are publications, directories, or partner sites that already cover competitors in this space. They're the highest-leverage outreach targets because they've already shown willingness to link to companies like Damco's.")
    parts.append("")
    if not analysis["intersection"]:
        parts.append("_None yet — needs backlink data for multiple competitors._")
    else:
        parts.append("| Referring Domain | # Competitors Linking | Total Links | Max Rank | Linked Competitors |")
        parts.append("|---|---:|---:|---:|---|")
        for r in analysis["intersection"][:30]:
            domains = ", ".join(f"`{d}`" for d in sorted(r["to_domains"] or [])[:5])
            parts.append(f"| `{r['source_domain']}` | {r['competitors_linked']} | "
                         f"{r['total_links']} | {r['max_rank'] or 0} | {domains} |")
        parts.append("")

    parts.append("## Top anchor text patterns across competitors")
    parts.append("")
    if not analysis["anchor_distribution"]:
        parts.append("_No anchor data yet._")
    else:
        parts.append("| Anchor Text | Uses | Competitors Using |")
        parts.append("|---|---:|---:|")
        for r in analysis["anchor_distribution"][:25]:
            parts.append(f"| `{(r['anchor'] or '')[:80]}` | {r['n']} | {r['competitors_using']} |")
        parts.append("")

    path.write_text("\n".join(parts), encoding="utf-8")
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(threat_tiers: tuple[str, ...] = DEFAULT_THREAT_TIERS,
        domain: str | None = None,
        limit: int = DEFAULT_LIMIT,
        cadence_days: int = DEFAULT_CADENCE_DAYS,
        force_all: bool = False,
        dry_run: bool = False,
        analyze_only: bool = False) -> dict:
    start = time.monotonic()
    today = date.today()

    # Decide what to fetch
    if analyze_only:
        targets = []
        logger.info("Analyze-only mode: skipping all API calls")
    else:
        targets = load_target_competitors(threat_tiers, domain, cadence_days, force_all)
        if not targets:
            logger.warning("No competitors due for backlink pull. "
                           "Use --all to force, or --analyze-only to report on existing data.")

    fetch_summary = {
        "competitors_targeted": len(targets),
        "successful_pulls":     0,
        "failed_pulls":         0,
        "total_fetched":        0,
        "total_inserted":       0,
        "access_denied":        False,
        "errors":               [],
    }

    # Fetch + store
    for i, comp in enumerate(targets, 1):
        logger.info("[%d/%d] Pulling backlinks for %s (limit=%d)...",
                    i, len(targets), comp["competitor_domain"], limit)
        counters = fetch_and_store(comp, limit, dry_run)
        fetch_summary["total_fetched"]  += counters["fetched"]
        fetch_summary["total_inserted"] += counters["inserted"]
        if counters.get("error"):
            fetch_summary["failed_pulls"] += 1
            if counters["error"] == "access_denied":
                fetch_summary["access_denied"] = True
                fetch_summary["errors"].append(
                    f"{comp['competitor_domain']}: access denied — Backlinks subscription required"
                )
                # No point hammering — stop here; one access-denied means all will fail
                logger.error("Backlinks API subscription required. Aborting further pulls.")
                logger.error("%s", counters.get("error_detail"))
                break
            else:
                fetch_summary["errors"].append(
                    f"{comp['competitor_domain']}: {counters.get('error_detail', counters['error'])}"
                )
        else:
            fetch_summary["successful_pulls"] += 1

    # Analysis runs against whatever's in the DB (whether fetched today or earlier)
    # Determine which competitor IDs to analyze
    if analyze_only:
        if domain:
            tier_filter = "AND c.competitor_domain = %s"
            params = [list(threat_tiers), domain]
        else:
            tier_filter = ""
            params = [list(threat_tiers)]
        rows = fetch_all(
            f"SELECT id FROM competitors c WHERE c.threat_tier = ANY(%s) AND c.is_tracked = TRUE {tier_filter}",
            params,
        )
        analyze_ids = [r["id"] for r in rows]
    else:
        analyze_ids = [t["id"] for t in targets if any(
            cb["competitor_id"] == t["id"]
            for cb in fetch_all(
                "SELECT competitor_id FROM competitor_backlinks WHERE competitor_id = %s LIMIT 1",
                [t["id"]],
            )
        )]
        # Also include any competitors not in `targets` that have stored data
        # (e.g. tier downgraded but old data still useful)
        rows = fetch_all(
            "SELECT DISTINCT competitor_id FROM competitor_backlinks "
            "WHERE competitor_id IN (SELECT id FROM competitors WHERE threat_tier = ANY(%s))",
            [list(threat_tiers)],
        )
        analyze_ids = list({r["competitor_id"] for r in rows} | set(analyze_ids))

    analysis = compute_analysis(analyze_ids)

    # Outputs
    xlsx_path = write_excel(analysis, competitors=targets)
    md_path   = write_markdown(analysis, fetch_summary)

    duration = time.monotonic() - start

    if not dry_run and not analyze_only:
        record_agent_run(
            agent_name=AGENT_NAME,
            status=("success" if fetch_summary["failed_pulls"] == 0
                    else "partial" if fetch_summary["successful_pulls"] > 0
                    else "error"),
            records_processed=fetch_summary["total_inserted"],
            errors=fetch_summary["errors"][:20],
            duration_seconds=round(duration, 2),
            metadata={
                "run_date":             today.isoformat(),
                "threat_tiers":         list(threat_tiers),
                "domain_filter":        domain,
                "limit":                limit,
                "cadence_days":         cadence_days,
                "force_all":            force_all,
                "competitors_targeted": fetch_summary["competitors_targeted"],
                "successful_pulls":     fetch_summary["successful_pulls"],
                "failed_pulls":         fetch_summary["failed_pulls"],
                "total_fetched":        fetch_summary["total_fetched"],
                "total_inserted":       fetch_summary["total_inserted"],
                "access_denied":        fetch_summary["access_denied"],
                "xlsx_path":            str(xlsx_path),
                "md_path":              str(md_path),
            },
        )

    # Console summary
    print()
    print(f"  {'=' * 72}")
    print(f"   BACKLINK ANALYZER — {today.isoformat()}")
    print(f"  {'=' * 72}")
    print()
    if fetch_summary["access_denied"]:
        print("  !! ACCESS DENIED to DataForSEO Backlinks API")
        print("  !! Subscription required: https://app.dataforseo.com/backlinks-subscription")
        print()
    print(f"  Competitors targeted:  {fetch_summary['competitors_targeted']}")
    print(f"  Successful pulls:      {fetch_summary['successful_pulls']}")
    print(f"  Failed pulls:          {fetch_summary['failed_pulls']}")
    print(f"  Backlinks fetched:     {fetch_summary['total_fetched']}")
    print(f"  Backlinks inserted:    {fetch_summary['total_inserted']}")
    print(f"  Competitors analyzed:  {len(analyze_ids)}")
    print(f"  Excel:                 {xlsx_path}")
    print(f"  Markdown:              {md_path}")
    print(f"  Duration:              {duration:.1f}s")
    print()

    return {
        "status":            ("error" if fetch_summary["access_denied"]
                              else "partial" if fetch_summary["failed_pulls"] > 0
                              else "success"),
        "total_fetched":     fetch_summary["total_fetched"],
        "total_inserted":    fetch_summary["total_inserted"],
        "access_denied":     fetch_summary["access_denied"],
        "xlsx_path":         str(xlsx_path),
        "md_path":           str(md_path),
        "duration_seconds":  round(duration, 2),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Damco Competitor Backlink Analyzer")
    parser.add_argument("--threat-tier", default=",".join(DEFAULT_THREAT_TIERS),
                        help=f"Comma-separated tiers (default: {','.join(DEFAULT_THREAT_TIERS)})")
    parser.add_argument("--domain", help="Pull only this specific competitor domain")
    parser.add_argument("--limit", type=int, default=DEFAULT_LIMIT,
                        help=f"Backlinks per competitor (default: {DEFAULT_LIMIT}, max: 1000)")
    parser.add_argument("--cadence", type=int, default=DEFAULT_CADENCE_DAYS,
                        help=f"Days between re-pulls per competitor (default: {DEFAULT_CADENCE_DAYS})")
    parser.add_argument("--all", dest="force_all", action="store_true",
                        help="Force re-pull ignoring cadence")
    parser.add_argument("--dry-run", action="store_true",
                        help="Call API (real cost) but don't write to DB")
    parser.add_argument("--analyze-only", action="store_true",
                        help="Skip API calls; analyze existing data and produce reports")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(name)s  %(levelname)s  %(message)s",
    )

    tiers = tuple(t.strip() for t in args.threat_tier.split(",") if t.strip())
    invalid = [t for t in tiers if t not in ("primary", "watch", "peripheral", "ignore")]
    if invalid:
        parser.error(f"Invalid threat tiers: {invalid}")

    if args.limit < 1 or args.limit > 1000:
        parser.error("--limit must be between 1 and 1000 (DataForSEO max)")

    run(threat_tiers=tiers, domain=args.domain, limit=args.limit,
        cadence_days=args.cadence, force_all=args.force_all,
        dry_run=args.dry_run, analyze_only=args.analyze_only)


if __name__ == "__main__":
    main()
