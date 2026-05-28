"""
Glossary Detector — Phase 1 module of Content Operations
=========================================================

Scans the active keyword list for "definition intent" phrases and
cross-references against existing glossary pages. Outputs a prioritized
list of *missing* glossary entries — ranked by search volume signal
(GSC impressions when available, otherwise raw keyword count).

Why glossary pages?
-------------------
Definition-intent searches ("what is X", "X meaning", "X definition")
are AEO/GEO gold. AI search engines (Google AI Overviews, Perplexity,
ChatGPT web search) overwhelmingly cite simple, well-structured
definitions. Damco's existing pages are largely service/marketing pages
that don't satisfy this intent — leaving a sizable visibility gap.

Detection patterns
------------------
Recognized definition-intent forms:
  "what is X"           -> term = X
  "X meaning"           -> term = X
  "X definition"        -> term = X
  "define X"            -> term = X
  "X explained"         -> term = X
  "X vs Y"              -> handled separately (comparison intent, future v2)
  "how does X work"     -> term = X
  "X for beginners"     -> term = X

A keyword that matches one of these patterns AND doesn't have a
corresponding glossary page becomes a "missing glossary" candidate.
A term already covered by an existing glossary page is dropped from
the recommendations.

Outputs
-------
- outputs/audits/glossary_gaps_<date>.md   (narrative)
- outputs/reports/glossary_gaps_<date>.xlsx (sortable data)
- Logs run to agent_runs

This module is intentionally rule-based — no LLM dependency. It's a
quick win that surfaces 50-200 high-value glossary opportunities from
the existing data without spending a dollar.

Usage
-----
    # Default: all active keywords, all offerings
    python -m content_operations.glossary_detector

    # One offering
    python -m content_operations.glossary_detector --offering "AI"

    # Lower the GSC-impression bar (show more candidates)
    python -m content_operations.glossary_detector --min-impressions 0

    # Dry run — analyze + write report, but skip agent_runs DB row
    python -m content_operations.glossary_detector --dry-run
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
import time
from collections import defaultdict
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from common.database import fetch_all, record_agent_run


logger = logging.getLogger("glossary_detector")
AGENT_NAME = "content_operations.glossary_detector"

OUTPUT_AUDITS  = Path(__file__).resolve().parent.parent / "outputs" / "audits"
OUTPUT_REPORTS = Path(__file__).resolve().parent.parent / "outputs" / "reports"

DEFAULT_MIN_IMPRESSIONS = 0   # 0 = show all; tune up to filter for high-value only


# ---------------------------------------------------------------------------
# Definition-intent patterns
# ---------------------------------------------------------------------------
# Each pattern is (regex, group-name-of-term, intent-label, signal-strength)
# signal-strength: 1.0 = unambiguous definition intent; lower = weaker
DEFINITION_PATTERNS: list[tuple[re.Pattern, str, str, float]] = [
    (re.compile(r"\bwhat\s+is\s+(?:an?\s+|the\s+)?(?P<term>.+?)(?:\?|$)",     re.IGNORECASE), "term", "what is X",         1.0),
    (re.compile(r"\bwhat\s+are\s+(?P<term>.+?)(?:\?|$)",                       re.IGNORECASE), "term", "what are X",        1.0),
    (re.compile(r"^(?P<term>.+?)\s+(?:meaning|definition)$",                   re.IGNORECASE), "term", "X meaning",          1.0),
    (re.compile(r"\bdefine\s+(?P<term>.+?)$",                                  re.IGNORECASE), "term", "define X",           1.0),
    (re.compile(r"^(?P<term>.+?)\s+explained$",                                re.IGNORECASE), "term", "X explained",        0.9),
    (re.compile(r"\bhow\s+does\s+(?P<term>.+?)\s+work(?:\?|$)",                re.IGNORECASE), "term", "how does X work",    0.9),
    (re.compile(r"^(?P<term>.+?)\s+for\s+beginners$",                          re.IGNORECASE), "term", "X for beginners",    0.7),
    (re.compile(r"^(?P<term>.+?)\s+guide$",                                    re.IGNORECASE), "term", "X guide",            0.6),
    (re.compile(r"^introduction\s+to\s+(?P<term>.+?)$",                        re.IGNORECASE), "term", "introduction to X",  0.8),
    (re.compile(r"^(?P<term>.+?)\s+(?:basics|fundamentals)$",                  re.IGNORECASE), "term", "X basics",           0.8),
]


# ---------------------------------------------------------------------------
# Read phase
# ---------------------------------------------------------------------------

def load_keywords(offering: str | None) -> list[dict]:
    """Pull every active keyword + its latest GSC stats."""
    params: list = []
    sql = """
        WITH latest_gsc AS (
            SELECT keyword_id, max(date) AS d FROM keyword_rankings
             WHERE source = 'gsc' GROUP BY keyword_id
        )
        SELECT k.id, k.keyword, k.offering, k.target_url,
               gsc.rank_position  AS gsc_position,
               gsc.clicks         AS gsc_clicks,
               gsc.impressions    AS gsc_impressions
          FROM keywords k
     LEFT JOIN latest_gsc lg ON lg.keyword_id = k.id
     LEFT JOIN keyword_rankings gsc ON gsc.keyword_id = k.id AND gsc.date = lg.d AND gsc.source = 'gsc'
         WHERE k.status = 'active'
    """
    if offering:
        sql += " AND k.offering = %s"
        params.append(offering)
    return fetch_all(sql, params)


def load_glossary_terms() -> set[str]:
    """
    Return the lowercased set of terms already covered by an existing
    glossary page. We extract the term from the URL path
    (everything after "/glossary/") plus any title we already audited.
    """
    rows = fetch_all(
        """
        SELECT url, title FROM pages
         WHERE page_type = 'glossary'
            OR url ILIKE %s
        """,
        ["%/glossary/%"],
    )
    covered: set[str] = set()
    for r in rows:
        url = (r.get("url") or "").lower()
        # /glossary/{term-slug}
        m = re.search(r"/glossary/([^/?#]+)", url)
        if m:
            term_slug = m.group(1).replace("-", " ").replace("_", " ").strip()
            if term_slug:
                covered.add(term_slug)
        # Also use title (without "What is" prefix and various suffixes)
        title = (r.get("title") or "").lower()
        if title:
            t = re.sub(r"^(what\s+(?:is|are)\s+|introduction\s+to\s+)", "", title)
            t = re.sub(r"\s+(?:meaning|definition|explained|guide)$", "", t)
            t = t.split("|")[0].split("–")[0].strip()
            if t:
                covered.add(t)
    return covered


# ---------------------------------------------------------------------------
# Detection
# ---------------------------------------------------------------------------

def extract_term(keyword: str) -> tuple[str | None, str, float]:
    """
    Run definition-intent patterns against a keyword. Returns
    (term_lowercased, pattern_label, signal_strength). All None if no match.
    """
    kw = (keyword or "").strip()
    for pattern, group_name, label, strength in DEFINITION_PATTERNS:
        m = pattern.search(kw)
        if m:
            term = m.group(group_name).strip().lower()
            # Trim filler words
            term = re.sub(r"^(an?\s+|the\s+)", "", term)
            if term and len(term) >= 2:
                return term, label, strength
    return None, "", 0.0


def detect_glossary_gaps(keywords: list[dict],
                        covered_terms: set[str],
                        min_impressions: int) -> list[dict]:
    """
    Group definition-intent keywords by their extracted term, then
    deduplicate and score each candidate.
    """
    by_term: dict[str, dict] = {}

    for kw in keywords:
        term, label, strength = extract_term(kw["keyword"])
        if not term:
            continue
        if term in covered_terms:
            # Already have a glossary page for this term — skip
            continue

        bucket = by_term.setdefault(term, {
            "term":              term,
            "matching_keywords": [],
            "max_strength":      0.0,
            "total_impressions": 0,
            "total_clicks":      0,
            "best_gsc_position": None,
            "offerings":         set(),
            "first_pattern":     label,
        })
        bucket["matching_keywords"].append({
            "keyword":      kw["keyword"],
            "offering":     kw["offering"],
            "pattern":      label,
            "strength":     strength,
            "impressions":  kw["gsc_impressions"] or 0,
            "clicks":       kw["gsc_clicks"] or 0,
            "gsc_position": kw["gsc_position"],
        })
        bucket["max_strength"]      = max(bucket["max_strength"], strength)
        bucket["total_impressions"] += kw["gsc_impressions"] or 0
        bucket["total_clicks"]      += kw["gsc_clicks"] or 0
        if kw["offering"]:
            bucket["offerings"].add(kw["offering"])
        # best_gsc_position = smallest (best) position seen across matching kws
        p = kw["gsc_position"]
        if p is not None:
            if bucket["best_gsc_position"] is None or p < bucket["best_gsc_position"]:
                bucket["best_gsc_position"] = p

    # Convert sets to lists; filter by impression threshold; score
    candidates: list[dict] = []
    for b in by_term.values():
        if b["total_impressions"] < min_impressions:
            continue
        b["offerings"] = sorted(b["offerings"])
        b["match_count"] = len(b["matching_keywords"])
        # Priority score: impressions are the strongest demand signal we have.
        # Tie-break on match_count (multiple matching phrasings = stronger signal).
        b["priority_score"] = round(
            b["max_strength"] * (
                (b["total_impressions"] / 100.0)
                + (b["total_clicks"] * 5)
                + (b["match_count"] * 2)
            ),
            2,
        )
        candidates.append(b)

    candidates.sort(key=lambda x: -x["priority_score"])
    return candidates


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def write_markdown(candidates: list[dict], offering: str | None,
                   covered_count: int, scanned_count: int) -> Path:
    OUTPUT_AUDITS.mkdir(parents=True, exist_ok=True)
    suffix = f"_{offering.replace(' ', '_').replace('/', '-')}" if offering else ""
    path = OUTPUT_AUDITS / f"glossary_gaps_{date.today().isoformat()}{suffix}.md"

    parts: list[str] = []
    parts.append(f"# Glossary Coverage Gaps — {date.today().isoformat()}")
    parts.append("")
    parts.append(f"_Generated by `{AGENT_NAME}`._")
    if offering:
        parts.append(f"_Scope: `{offering}`_")
    parts.append("")

    parts.append("## Summary")
    parts.append("")
    parts.append("| Metric | Value |")
    parts.append("|---|---:|")
    parts.append(f"| Active keywords scanned | {scanned_count} |")
    parts.append(f"| Existing glossary terms (excluded) | {covered_count} |")
    parts.append(f"| Definition-intent gaps surfaced | **{len(candidates)}** |")
    if candidates:
        with_impr = sum(1 for c in candidates if c["total_impressions"] > 0)
        with_clicks = sum(1 for c in candidates if c["total_clicks"] > 0)
        parts.append(f"| Gaps with GSC impressions (real demand) | {with_impr} |")
        parts.append(f"| Gaps with GSC clicks (people landing on something) | {with_clicks} |")
    parts.append("")

    if not candidates:
        parts.append("_No missing glossary terms detected. Either coverage is complete or the patterns didn't match any active keyword._")
        path.write_text("\n".join(parts), encoding="utf-8")
        return path

    # Top 30 prioritized
    parts.append("## Top 30 prioritized gaps")
    parts.append("")
    parts.append("Sorted by priority score (signal strength × impressions × clicks × match count).")
    parts.append("")
    parts.append("| # | Term | Priority | Matches | GSC Impressions (14d) | GSC Clicks (14d) | Best Pos | Offerings |")
    parts.append("|---:|---|---:|---:|---:|---:|---|---|")
    for i, c in enumerate(candidates[:30], 1):
        offerings = ", ".join(c["offerings"]) or "—"
        best_pos = f"{c['best_gsc_position']:.1f}" if c["best_gsc_position"] else "—"
        parts.append(f"| {i} | `{c['term']}` | {c['priority_score']} | {c['match_count']} | "
                     f"{c['total_impressions']} | {c['total_clicks']} | {best_pos} | {offerings} |")
    parts.append("")

    # Detail for the top 10 — show the matching keywords
    parts.append("## Detail — top 10 candidates with matching keywords")
    parts.append("")
    for i, c in enumerate(candidates[:10], 1):
        parts.append(f"### {i}. `{c['term']}`")
        parts.append("")
        parts.append(f"- **Priority score:** {c['priority_score']}")
        parts.append(f"- **Matching keywords:** {c['match_count']}")
        parts.append(f"- **Offerings:** {', '.join(c['offerings']) or '—'}")
        parts.append(f"- **GSC impressions (14d):** {c['total_impressions']}")
        parts.append(f"- **GSC clicks (14d):** {c['total_clicks']}")
        parts.append("")
        parts.append("| Matching keyword | Pattern | Impr | Clicks | GSC Pos | Offering |")
        parts.append("|---|---|---:|---:|---:|---|")
        for mk in c["matching_keywords"][:10]:
            pos = f"{mk['gsc_position']:.1f}" if mk["gsc_position"] else "—"
            parts.append(f"| `{mk['keyword']}` | {mk['pattern']} | {mk['impressions']} | "
                         f"{mk['clicks']} | {pos} | {mk['offering'] or '—'} |")
        parts.append("")

    path.write_text("\n".join(parts), encoding="utf-8")
    return path


def write_excel(candidates: list[dict]) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    OUTPUT_REPORTS.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_REPORTS / f"glossary_gaps_{date.today().isoformat()}.xlsx"
    wb = openpyxl.Workbook()
    fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    bold = Font(bold=True, color="FFFFFF")

    def header(ws, headers):
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = bold
            c.fill = fill

    # Sheet 1: ranked candidates
    ws = wb.active
    ws.title = "Glossary Gaps"
    header(ws, ["Rank", "Term", "Priority Score", "Match Count",
                "Total Impressions (14d)", "Total Clicks (14d)",
                "Best GSC Position", "Offerings", "Strongest Pattern"])
    for i, c in enumerate(candidates, 1):
        ws.append([
            i, c["term"], c["priority_score"], c["match_count"],
            c["total_impressions"], c["total_clicks"],
            float(c["best_gsc_position"]) if c["best_gsc_position"] else "",
            ", ".join(c["offerings"]), c["first_pattern"],
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col, w in zip("ABCDEFGHI", [6, 42, 12, 12, 22, 18, 16, 36, 22]):
        ws.column_dimensions[col].width = w

    # Sheet 2: matching keywords detail (long format)
    ws2 = wb.create_sheet("Matching Keywords")
    header(ws2, ["Term (Gap)", "Keyword", "Pattern", "Impressions",
                 "Clicks", "GSC Position", "Offering"])
    for c in candidates:
        for mk in c["matching_keywords"]:
            ws2.append([
                c["term"], mk["keyword"], mk["pattern"],
                mk["impressions"], mk["clicks"],
                float(mk["gsc_position"]) if mk["gsc_position"] else "",
                mk["offering"] or "",
            ])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    for col, w in zip("ABCDEFG", [38, 48, 22, 14, 12, 14, 36]):
        ws2.column_dimensions[col].width = w

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(offering: str | None = None,
        min_impressions: int = DEFAULT_MIN_IMPRESSIONS,
        dry_run: bool = False) -> dict:
    start = time.monotonic()

    keywords = load_keywords(offering)
    if not keywords:
        logger.warning("No active keywords found%s", f" for offering={offering}" if offering else "")
        return {"status": "skipped", "reason": "no keywords"}

    covered = load_glossary_terms()
    logger.info("Scanning %d active keyword(s); %d existing glossary term(s) excluded",
                len(keywords), len(covered))

    candidates = detect_glossary_gaps(keywords, covered, min_impressions)
    logger.info("Surfaced %d glossary gap candidates", len(candidates))

    md_path   = write_markdown(candidates, offering, len(covered), len(keywords))
    xlsx_path = write_excel(candidates)

    duration = time.monotonic() - start

    if not dry_run:
        record_agent_run(
            agent_name=AGENT_NAME,
            status="success",
            records_processed=len(candidates),
            errors=[],
            duration_seconds=round(duration, 2),
            metadata={
                "run_date":          date.today().isoformat(),
                "offering":          offering,
                "min_impressions":   min_impressions,
                "keywords_scanned":  len(keywords),
                "existing_glossary": len(covered),
                "candidates":        len(candidates),
                "with_impressions":  sum(1 for c in candidates if c["total_impressions"] > 0),
                "with_clicks":       sum(1 for c in candidates if c["total_clicks"] > 0),
                "md_path":           str(md_path),
                "xlsx_path":         str(xlsx_path),
            },
        )

    # Console summary
    print()
    print(f"  {'=' * 72}")
    print(f"   GLOSSARY DETECTOR -- {date.today().isoformat()}")
    print(f"  {'=' * 72}")
    print()
    print(f"  Active keywords scanned:    {len(keywords)}")
    print(f"  Existing glossary terms:    {len(covered)}")
    print(f"  Gap candidates surfaced:    {len(candidates)}")
    print(f"  With GSC impressions:       {sum(1 for c in candidates if c['total_impressions'] > 0)}")
    print(f"  With GSC clicks:            {sum(1 for c in candidates if c['total_clicks'] > 0)}")
    print(f"  Markdown:                   {md_path}")
    print(f"  Excel:                      {xlsx_path}")
    print(f"  Duration:                   {duration:.2f}s")
    print()

    if candidates[:5]:
        print(f"  Top 5 candidates:")
        for c in candidates[:5]:
            offerings = ", ".join(c["offerings"]) or "-"
            print(f"    {c['term'][:50]:<52}  pri={c['priority_score']:>6.1f}  "
                  f"impr={c['total_impressions']:>5}  clicks={c['total_clicks']:>3}  "
                  f"[{offerings}]")
        print()

    return {
        "status":           "success",
        "candidates":       len(candidates),
        "md_path":          str(md_path),
        "xlsx_path":        str(xlsx_path),
        "duration_seconds": round(duration, 2),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Damco Glossary Gap Detector")
    parser.add_argument("--offering", help="Restrict to one offering (default: all)")
    parser.add_argument("--min-impressions", type=int, default=DEFAULT_MIN_IMPRESSIONS,
                        help=f"Only show candidates with at least N GSC impressions "
                             f"(default: {DEFAULT_MIN_IMPRESSIONS})")
    parser.add_argument("--dry-run", action="store_true",
                        help="Generate report but skip agent_runs DB row")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(name)s  %(levelname)s  %(message)s",
    )

    run(offering=args.offering, min_impressions=args.min_impressions, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
