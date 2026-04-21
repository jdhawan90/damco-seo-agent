"""
Import BPM (Business Process Management) keywords from a mastersheet.

The BPO mastersheet has a different structure from the main ranking Excel:
  - Single "Keywords" sheet
  - Columns: Category, Page Priority, Keywords Strategy, SEO Member,
             Keywords, SEO Executive
  - Category is treated as the sub-offering (services). Offering is
    set to "BPM" for all keywords.

Usage
-----
    python -m keyword_intelligence.import_bpm_keywords <path-to-excel> \
        [--executive NAME] [--offering NAME] [--sheet NAME]

    # Default: all keywords assigned to Abhishek under "BPM" offering
    python -m keyword_intelligence.import_bpm_keywords "bpm.xlsx"
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from common.database import connection, fetch_all


COLUMN_HEADERS = {
    "category": "Category",
    "priority": "Page Priority",
    "strategy": "Keywords Strategy",
    "member": "SEO Member",
    "keyword": "Keywords",
    "executive": "SEO Executive",
}


def find_columns(ws) -> dict[str, int]:
    """Map logical column name → 1-based column index by matching header row."""
    resolved: dict[str, int] = {}
    headers = {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}
    for key, header_name in COLUMN_HEADERS.items():
        if header_name in headers:
            resolved[key] = headers[header_name]
    return resolved


def extract_keywords(filepath: str, sheet_name: str) -> list[dict]:
    """Extract all keywords from the mastersheet, preserving category fill-down."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    cols = find_columns(ws)

    missing = [k for k in ("keyword", "category") if k not in cols]
    if missing:
        raise RuntimeError(f"Missing required columns: {missing}")

    print(f"  Columns resolved: {cols}")

    keywords: list[dict] = []
    last_category = None

    for row_idx in range(2, ws.max_row + 1):
        cat_val = ws.cell(row_idx, cols["category"]).value if "category" in cols else None
        kw_val = ws.cell(row_idx, cols["keyword"]).value
        priority = ws.cell(row_idx, cols["priority"]).value if "priority" in cols else None
        strategy = ws.cell(row_idx, cols["strategy"]).value if "strategy" in cols else None
        exec_val = ws.cell(row_idx, cols["executive"]).value if "executive" in cols else None

        # Fill-down for merged category cells
        if cat_val:
            last_category = str(cat_val).strip()

        if not kw_val or not str(kw_val).strip():
            continue

        priority_str = str(priority).strip().lower() if priority else "high"  # default
        if priority_str not in ("high", "medium", "low"):
            priority_str = "high"

        strategy_str = str(strategy).strip() if strategy else None
        if strategy_str:
            # Normalize — sheet has trailing spaces on "Primary " etc.
            strategy_str = strategy_str.strip().lower()

        keywords.append({
            "keyword": str(kw_val).strip().lower(),
            "services": last_category,
            "page_priority": priority_str,
            "strategy": strategy_str,
            "executive": str(exec_val).strip() if exec_val else None,
        })

    return keywords


def import_to_db(
    keywords: list[dict],
    offering: str,
    override_executive: str | None,
    sheet_label: str,
) -> dict:
    """Import keywords, assigning them to the given offering and executive."""

    # --- 1. Upsert the target executive ---
    target_exec = override_executive
    if not target_exec and keywords:
        # Use whatever's in the sheet (should be consistent)
        execs = {kw["executive"] for kw in keywords if kw["executive"]}
        if len(execs) == 1:
            target_exec = next(iter(execs))

    if not target_exec:
        raise RuntimeError("No executive specified — use --executive")

    exec_id = None
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO seo_executives (name)
                VALUES (%s)
                ON CONFLICT (name) DO UPDATE SET name = EXCLUDED.name
                RETURNING id
            """, (target_exec,))
            exec_id = cur.fetchone()[0]

    print(f"  Executive: {target_exec} (id={exec_id})")

    # --- 2. Insert keywords (additive — no wipe) ---
    inserted = 0
    updated = 0
    assigned = 0
    kw_ids: list[int] = []

    with connection() as conn:
        with conn.cursor() as cur:
            for kw in keywords:
                cur.execute("""
                    INSERT INTO keywords
                        (keyword, offering, services, importance, type, status)
                    VALUES (%s, %s, %s, %s, 'primary', 'active')
                    ON CONFLICT (keyword, offering) DO UPDATE SET
                        services   = COALESCE(EXCLUDED.services, keywords.services),
                        importance = EXCLUDED.importance
                    RETURNING id, xmax = 0 AS is_insert
                """, (
                    kw["keyword"],
                    offering,
                    kw["services"],
                    kw["page_priority"],
                ))
                row = cur.fetchone()
                kw_id = row[0]
                if row[1]:
                    inserted += 1
                else:
                    updated += 1
                kw_ids.append(kw_id)

                # Assign to executive
                cur.execute("""
                    INSERT INTO executive_keyword_assignments
                        (executive_id, keyword_id, sheet_source)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (executive_id, keyword_id) DO NOTHING
                    RETURNING id
                """, (exec_id, kw_id, sheet_label))
                if cur.fetchone():
                    assigned += 1

    return {
        "inserted": inserted,
        "updated": updated,
        "assigned": assigned,
        "executive": target_exec,
        "offering": offering,
        "total": len(keywords),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Import BPM keywords from mastersheet")
    parser.add_argument("filepath", help="Path to the BPM Excel file")
    parser.add_argument("--sheet", default="Keywords", help="Sheet name (default: Keywords)")
    parser.add_argument("--offering", default="BPM",
                        help="Offering name to assign (default: BPM)")
    parser.add_argument("--executive", default="Abhishek",
                        help="Executive to assign keywords to (default: Abhishek)")
    args = parser.parse_args()

    if not Path(args.filepath).exists():
        print(f"ERROR: File not found: {args.filepath}")
        sys.exit(1)

    print(f"\n  Source: {args.filepath}")
    print(f"  Sheet:  {args.sheet}")

    keywords = extract_keywords(args.filepath, args.sheet)
    print(f"  Extracted {len(keywords)} keywords")

    if not keywords:
        print("  Nothing to import. Aborting.")
        sys.exit(1)

    # Summarize
    categories: dict[str, int] = {}
    priorities: dict[str, int] = {}
    for kw in keywords:
        cat = kw["services"] or "(none)"
        categories[cat] = categories.get(cat, 0) + 1
        pri = kw["page_priority"]
        priorities[pri] = priorities.get(pri, 0) + 1

    print(f"\n  Categories ({len(categories)}):")
    for cat, n in sorted(categories.items(), key=lambda x: -x[1]):
        print(f"    {cat:<45s} {n:>3}")

    print(f"\n  Page priorities: {priorities}")

    # Import
    print()
    result = import_to_db(
        keywords=keywords,
        offering=args.offering,
        override_executive=args.executive,
        sheet_label=args.sheet,
    )

    print("\n  === Import Summary ===")
    print(f"    Offering:         {result['offering']}")
    print(f"    Executive:        {result['executive']}")
    print(f"    Keywords total:   {result['total']}")
    print(f"    Newly inserted:   {result['inserted']}")
    print(f"    Updated existing: {result['updated']}")
    print(f"    New assignments:  {result['assigned']}")

    # Verification
    verify = fetch_all("""
        SELECT count(*) as n FROM executive_keyword_assignments a
        JOIN seo_executives e ON e.id = a.executive_id
        JOIN keywords k ON k.id = a.keyword_id
        WHERE e.name = %s AND k.offering = %s
    """, [result["executive"], result["offering"]])
    print(f"\n    Total {result['executive']} -> {result['offering']}: {verify[0]['n']} keywords")
    print()


if __name__ == "__main__":
    main()
