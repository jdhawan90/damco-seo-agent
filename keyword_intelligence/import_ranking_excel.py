"""
Import keywords from the master ranking Excel file.

Reads 'Service Pages' and 'Tech Pages' sheets, extracts keywords marked
as importance=High, and imports them into the database with executive
assignment tagging.

Usage
-----
    python -m keyword_intelligence.import_ranking_excel <path-to-excel>

    # With --wipe flag to clear existing data first
    python -m keyword_intelligence.import_ranking_excel <path-to-excel> --wipe
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from common.database import connection, fetch_all


# Sheets to process and their column mappings
SHEET_CONFIG = {
    "Service Pages": {
        "offering_col": "A",
        "services_col": "B",
        "keyword_col": "C",
        "google_sv_col": "D",
        "semrush_sv_col": "E",
        "imp_header": "Imp.",
        "exec_header": "SEO Executive",
    },
    "Tech Pages": {
        "offering_col": "A",
        "services_col": "B",
        "keyword_col": "C",
        "google_sv_col": "D",
        "semrush_sv_col": "E",
        "imp_header": "Imp.",
        "exec_header": "SEO Executive",
    },
}


def find_column_by_header(ws, header_name: str) -> str | None:
    """Find column letter by matching header row text."""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().replace("\n", " ") == header_name:
            return cell.column_letter
    return None


def get_cell_by_letter(row, col_letter: str):
    """Get cell value from a row by column letter."""
    for cell in row:
        if cell.column_letter == col_letter:
            return cell.value
    return None


def extract_keywords(filepath: str) -> list[dict]:
    """
    Extract high-importance keywords from both sheets.
    Handles merged-cell inheritance (offering/services propagate down).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_keywords: list[dict] = []

    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]

        # Find dynamic columns (Imp. and SEO Executive) by header
        imp_col = find_column_by_header(ws, config["imp_header"])
        exec_col = find_column_by_header(ws, config["exec_header"])

        if not imp_col:
            print(f"  WARNING: '{config['imp_header']}' column not found in '{sheet_name}'")
            continue

        print(f"  {sheet_name}: Imp.={imp_col}, Executive={exec_col or 'N/A'}")

        # Track last seen offering/services for merged-cell fill-down
        last_offering = None
        last_services = None

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            offering = get_cell_by_letter(row, config["offering_col"])
            services = get_cell_by_letter(row, config["services_col"])
            keyword = get_cell_by_letter(row, config["keyword_col"])
            google_sv = get_cell_by_letter(row, config["google_sv_col"])
            semrush_sv = get_cell_by_letter(row, config["semrush_sv_col"])
            imp = get_cell_by_letter(row, imp_col)
            executive = get_cell_by_letter(row, exec_col) if exec_col else None

            # Fill-down for merged cells
            if offering:
                last_offering = str(offering).strip()
            if services:
                last_services = str(services).strip()

            # Skip if no keyword
            if not keyword or not str(keyword).strip():
                continue

            keyword_text = str(keyword).strip().lower()

            # Only import high-importance keywords
            if str(imp).strip().lower() != "high":
                continue

            # Normalize executive name
            exec_name = None
            if executive:
                exec_name = str(executive).strip()
                # Title-case normalization (e.g., "khushbu" → "Khushbu")
                exec_name = exec_name.title()

            # Parse search volume
            sv_text = str(google_sv).strip() if google_sv else None
            sv_numeric = None
            if semrush_sv and isinstance(semrush_sv, (int, float)):
                sv_numeric = int(semrush_sv)

            all_keywords.append({
                "keyword": keyword_text,
                "offering": last_offering,
                "services": last_services,
                "google_sv": sv_text,
                "semrush_sv": sv_numeric,
                "importance": "high",
                "executive": exec_name,
                "sheet": sheet_name,
            })

    print(f"\n  Total high-importance keywords extracted: {len(all_keywords)}")
    return all_keywords


def wipe_existing_data() -> None:
    """Delete all existing keyword-related data."""
    print("\n  Wiping existing data...")
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM executive_keyword_assignments")
            cur.execute("DELETE FROM keyword_rankings")
            cur.execute("DELETE FROM keyword_search_volume")
            cur.execute("DELETE FROM keywords")
            cur.execute("DELETE FROM seo_executives")
            print("  Cleared: executive_keyword_assignments, keyword_rankings,")
            print("           keyword_search_volume, keywords, seo_executives")


def import_keywords(keywords: list[dict]) -> None:
    """Import extracted keywords into the database."""

    # --- 1. Collect unique executives ---
    executives = sorted({kw["executive"] for kw in keywords if kw["executive"]})
    print(f"\n  Executives: {executives}")

    exec_id_map: dict[str, int] = {}
    with connection() as conn:
        with conn.cursor() as cur:
            for name in executives:
                cur.execute("""
                    INSERT INTO seo_executives (name)
                    VALUES (%s)
                    ON CONFLICT (name) DO UPDATE SET name = EXCLUDED.name
                    RETURNING id
                """, (name,))
                exec_id_map[name] = cur.fetchone()[0]

    print(f"  Executives upserted: {len(exec_id_map)}")

    # --- 2. Insert keywords (deduplicate by keyword + offering) ---
    kw_id_map: dict[tuple[str, str], int] = {}
    inserted = 0
    skipped = 0

    with connection() as conn:
        with conn.cursor() as cur:
            for kw in keywords:
                key = (kw["keyword"], kw["offering"] or "")
                if key in kw_id_map:
                    skipped += 1
                    continue

                cur.execute("""
                    INSERT INTO keywords
                        (keyword, offering, services, google_sv, importance, type, status)
                    VALUES (%s, %s, %s, %s, %s, 'primary', 'active')
                    ON CONFLICT (keyword, offering) DO UPDATE SET
                        services   = COALESCE(EXCLUDED.services, keywords.services),
                        google_sv  = COALESCE(EXCLUDED.google_sv, keywords.google_sv),
                        importance = EXCLUDED.importance
                    RETURNING id
                """, (
                    kw["keyword"],
                    kw["offering"],
                    kw["services"],
                    kw["google_sv"],
                    kw["importance"],
                ))
                kw_id_map[key] = cur.fetchone()[0]
                inserted += 1

    print(f"  Keywords inserted: {inserted}, duplicates skipped: {skipped}")

    # --- 3. Also store semrush search volume where available ---
    sv_count = 0
    with connection() as conn:
        with conn.cursor() as cur:
            for kw in keywords:
                if kw["semrush_sv"] and kw["semrush_sv"] > 0:
                    key = (kw["keyword"], kw["offering"] or "")
                    kw_id = kw_id_map.get(key)
                    if kw_id:
                        cur.execute("""
                            INSERT INTO keyword_search_volume
                                (keyword_id, date, search_volume, source)
                            VALUES (%s, CURRENT_DATE, %s, 'manual')
                            ON CONFLICT (keyword_id, date, source) DO NOTHING
                        """, (kw_id, kw["semrush_sv"]))
                        sv_count += 1

    print(f"  Search volumes stored: {sv_count}")

    # --- 4. Create executive-keyword assignments ---
    assign_count = 0
    with connection() as conn:
        with conn.cursor() as cur:
            for kw in keywords:
                if not kw["executive"]:
                    continue
                key = (kw["keyword"], kw["offering"] or "")
                kw_id = kw_id_map.get(key)
                exec_id = exec_id_map.get(kw["executive"])
                if kw_id and exec_id:
                    cur.execute("""
                        INSERT INTO executive_keyword_assignments
                            (executive_id, keyword_id, sheet_source)
                        VALUES (%s, %s, %s)
                        ON CONFLICT (executive_id, keyword_id) DO NOTHING
                    """, (exec_id, kw_id, kw["sheet"]))
                    assign_count += 1

    print(f"  Executive assignments created: {assign_count}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import high-importance keywords from the master ranking Excel"
    )
    parser.add_argument("filepath", help="Path to the ranking Excel file")
    parser.add_argument("--wipe", action="store_true",
                        help="Delete all existing keyword data before importing")
    args = parser.parse_args()

    if not Path(args.filepath).exists():
        print(f"ERROR: File not found: {args.filepath}")
        sys.exit(1)

    print(f"\n  Source: {args.filepath}")

    keywords = extract_keywords(args.filepath)
    if not keywords:
        print("  No high-importance keywords found. Aborting.")
        sys.exit(1)

    if args.wipe:
        wipe_existing_data()

    import_keywords(keywords)

    # Summary
    print("\n  === Import Summary ===")
    for r in fetch_all("""
        SELECT e.name, count(a.id) as keywords
        FROM seo_executives e
        LEFT JOIN executive_keyword_assignments a ON a.executive_id = e.id
        GROUP BY e.name ORDER BY e.name
    """):
        print(f"    {r['name']:<20s} {r['keywords']:>4} keywords")

    total = fetch_all("SELECT count(*) as n FROM keywords WHERE importance = 'high'")[0]["n"]
    print(f"\n  Total high-importance keywords in DB: {total}")
    print()


if __name__ == "__main__":
    main()
