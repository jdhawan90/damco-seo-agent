"""
Trend Scout — emerging-keyword discovery
========================================

Finds the phrases the industry has started using that Damco isn't tracking
yet, and proposes them as keywords with real search-volume evidence.

Standard agent lifecycle:
  Read    — poll the feed registry (`trend_sources`) via common.connectors.feeds:
            tech press (CIO.com, Techmeme, InfoWorld), practitioner communities
            (Reddit, Hacker News), and blogging platforms (Medium tag feeds).
  Process — extract candidate n-grams, drop the ones we already track, map each
            to an offering (rule-based first, Claude only for the residue),
            fetch Google Ads Keyword Planner volume + 12-month trend, and score.
  Write   — upsert `trend_mentions` (the evidence trail) and `keyword_candidates`
            (the deliverable). Nothing enters `keywords` automatically.
  Notify  — console summary, Excel workbook, and a markdown digest.

Design notes
------------
*Rule-based first, LLM second.* N-gram extraction, novelty checking, volume
lookup, and scoring are all deterministic. Claude is used for exactly one
thing: turning a raw buzz phrase into a search-shaped keyword and assigning
it an offering, and only for candidates the token rules couldn't classify.
The whole pipeline degrades to rule-only output if the API key is missing.

*Discovery is cheap; tracking is expensive.* Every keyword in `keywords`
costs ~$0.00465 on every rank-tracker run. So candidates live in their own
table and reach the tracked set only when a human runs `--promote`.

Usage
-----
    # Full discovery run: harvest, score, write, report
    python -m keyword_intelligence.trend_scout

    # Look back further than the default 14 days
    python -m keyword_intelligence.trend_scout --days 30

    # Only sources hinted at one offering
    python -m keyword_intelligence.trend_scout --offering AI

    # Harvest and score, but skip the paid volume lookup
    python -m keyword_intelligence.trend_scout --no-volume

    # Skip Claude; pure token-rule classification
    python -m keyword_intelligence.trend_scout --no-llm

    # See what it would do without writing anything
    python -m keyword_intelligence.trend_scout --dry-run

    # Review queue
    python -m keyword_intelligence.trend_scout --list-candidates --min-score 45

    # Human gate: move approved candidates into the tracked keyword set
    python -m keyword_intelligence.trend_scout --promote --ids 12,15,31
"""

from __future__ import annotations

import argparse
import functools
import hashlib
import json
import logging
import math
import re
import sys
import time
from collections import defaultdict
from datetime import date, datetime, timezone
from pathlib import Path

import psycopg2.extras

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from common.config import settings
from common.connectors import feeds
from common.connectors.dataforseo import (
    COST_PER_KEYWORD_PLANNER_TASK,
    KEYWORD_PLANNER_BATCH_SIZE,
    DataForSEOError,
    get_search_volume,
)
from common.database import connection, fetch_all, record_agent_run
from common.llm import LLMUnavailableError, call_claude
from common.tenant import profile, system_preamble


logger = logging.getLogger(__name__)

AGENT_NAME = "keyword_intelligence.trend_scout"

# ---------------------------------------------------------------------------
# Extraction tuning
# ---------------------------------------------------------------------------

DEFAULT_LOOKBACK_DAYS = 14
DEFAULT_MAX_ITEMS_PER_SOURCE = 60
# Below this many mentions a phrase is one person's turn of phrase, not a trend.
MIN_MENTIONS = 2
# A phrase seen in one outlet is that outlet's house style. Two is a signal.
MIN_SOURCE_SPREAD = 2
# How many mentions from offering-hinted sources it takes before the hint is
# allowed to classify a phrase the token rules couldn't place.
MIN_HINT_SUPPORT = 2
NGRAM_RANGE = (2, 5)
# Cap the volume lookup so a noisy harvest can't run up a bill.
DEFAULT_MAX_VOLUME_LOOKUPS = 600
# Candidates whose token overlap with a tracked keyword exceeds this are
# treated as duplicates of what we already track.
NOVELTY_SIMILARITY_THRESHOLD = 0.80
# Claude classifies at most this many unresolved phrases per run.
MAX_LLM_CANDIDATES = 150

# Scoring weights — sum to 100.
SCORE_WEIGHTS: dict[str, int] = {
    "buzz":        30,   # how much the industry is talking about it
    "volume":      25,   # how many people search it
    "momentum":    20,   # whether that search demand is rising
    "opportunity": 15,   # whether it's genuinely new territory for us
    "commercial":  10,   # whether the traffic would be worth anything
}
assert sum(SCORE_WEIGHTS.values()) == 100

# Category diversity multiplier: a phrase appearing in tech press *and*
# practitioner communities is a stronger signal than one confined to either.
CATEGORY_SPREAD_BONUS = {1: 1.00, 2: 1.25, 3: 1.45, 4: 1.60}


# ---------------------------------------------------------------------------
# Vocabulary
# ---------------------------------------------------------------------------

# Ordinary English + newsroom filler. A phrase made only of these is noise.
STOPWORDS: frozenset[str] = frozenset("""
a about above after again against all also am an and any are aren as at be
because been before being below between both but by can cannot could couldn
did didn do does doesn doing don down during each few for from further had
hadn has hasn have haven having he her here hers herself him himself his how
i if in into is isn it its itself just let me more most must my myself no nor
not now of off on once only or other ought our ours ourselves out over own
same shan she should shouldn so some such than that the their theirs them
themselves then there these they this those through to too under until up
very was wasn we were weren what when where which while who whom why will
with won would wouldn you your yours yourself yourselves
said says say new news today yesterday tomorrow week month year years time
first last next best top good great big small make makes made using use used
get gets getting going go goes went one two three four five ten
via according report reports reported reporting according update updates
via amp nbsp read comment comments post posts article articles blog blogs
things thing way ways lot lots need needs want wants like likes look looks
much many may might see seen know knows think thinks people company companies
""".split())

# Prose connectives that mark a sentence fragment rather than a noun phrase.
# "ai is transforming", "faster with ai", "including ai" are how journalists
# write; nobody types them into Google.
#
# Deliberately excludes and/or/of/for/in/on/with/to/as/by/from — those appear
# in perfectly good keywords ("application support and maintenance services",
# "crm for insurance", "migration to azure").
PROSE_MARKERS: frozenset[str] = frozenset("""
is are was were be been being am has have had having
will would shall should can could may might must
said says say told tells adds noted claims argues
including according while when where why how because however
although though whether unless since during despite amid
this that these those which who whom whose
they them we us you your our their its his her he she it
what who's don't doesn't isn't aren't
to by per with into onto upon over under after before against
between through across toward towards within without about
next last first back out up off down again else ago
frequently enough quickly rapidly recently currently already
finally simply actually basically literally really very quite
rather fairly extremely increasingly significantly potentially
essentially effectively particularly especially generally
typically usually often sometimes always never still yet also too
""".split())

# The tenant-specific half of the vocabulary — news nouns, commercial
# tokens, and the offering token map — lives in the tenant profile.
#
# Wrapped rather than read inline for two reasons: importing this module
# must not touch the database, and `offering_matchers` / `offering_marker_words`
# rebuild their collections on every property access while _is_viable_phrase
# runs once per extracted n-gram — tens of thousands of times a run.

@functools.lru_cache(maxsize=1)
def _commercial_tokens() -> frozenset[str]:
    """Tokens that make a phrase read like something someone would search."""
    return profile().vocab("commercial_tokens")


@functools.lru_cache(maxsize=1)
def _generic_heads() -> frozenset[str]:
    """News nouns. Offering marker + one of these is a headline, not a query."""
    return profile().vocab("generic_heads")


@functools.lru_cache(maxsize=1)
def _offering_matchers() -> tuple[tuple[str, str], ...]:
    """(token, offering) longest-first, so "power bi" outranks "bi"."""
    return profile().offering_matchers


@functools.lru_cache(maxsize=1)
def _offering_marker_words() -> frozenset[str]:
    """Every individual word appearing in any offering token."""
    return profile().offering_marker_words


# Two chars minimum: a one-letter token is almost always the orphaned half of
# a split possessive ("Google's" -> "google", "s"), which produced phrases
# like "s next frontier ai model" in the first live run.
_TOKEN_RE = re.compile(r"[a-z0-9][a-z0-9+#./&'-]+")
_WS_RE = re.compile(r"\s+")


# ---------------------------------------------------------------------------
# Console safety — Windows cp1252 chokes on feed unicode
# ---------------------------------------------------------------------------

def _safe_console(s: str | None) -> str:
    """
    Make an arbitrary string printable on a Windows cp1252 console.

    Feed titles are full of typographic unicode (curly quotes, em dashes,
    narrow no-break spaces, zero-width joiners). Printing them raw kills the
    run *after* the DB writes have committed — the worst possible failure
    mode. Same three-pass approach the rank tracker uses.
    """
    if not s:
        return ""
    # Pass 1: visible-whitespace unicode -> regular space (must not be deleted,
    # or "machine learning" becomes "machinelearning").
    for ch in ("\u00a0", "\u202f", "\u2007", "\u2008", "\u2009", "\u200a",
               "\u205f", "\u3000"):
        s = s.replace(ch, " ")
    # Pass 2: truly invisible joiners -> delete.
    for ch in ("\u200b", "\u200c", "\u200d", "\u2060", "\ufeff", "\u180e"):
        s = s.replace(ch, "")
    # Pass 3: encode-replace backstop for everything else.
    enc = (sys.stdout.encoding or "utf-8")
    try:
        return s.encode(enc, errors="replace").decode(enc, errors="replace")
    except (LookupError, UnicodeError):
        return s.encode("ascii", errors="replace").decode("ascii")


# ---------------------------------------------------------------------------
# Read — the feed registry
# ---------------------------------------------------------------------------

def load_sources(offering: str | None = None, include_disabled: bool = False) -> list[dict]:
    """Fetch the feed registry, optionally narrowed to one offering hint."""
    sql = "SELECT * FROM trend_sources WHERE TRUE"
    params: list = []
    if not include_disabled:
        sql += " AND enabled"
    if offering:
        sql += " AND offering_hint = %s"
        params.append(offering)
    sql += " ORDER BY category, name"
    return fetch_all(sql, params)


# A source that fails this many runs in a row is probably dead, not unlucky.
UNHEALTHY_FAILURE_THRESHOLD = 3
RETIRE_FAILURE_THRESHOLD = 10


def warn_unhealthy_sources() -> list[dict]:
    """
    Surface feeds that have been failing run after run.

    A silently-dead source is worse than a loud one: the harvest still
    "succeeds", just with a blind spot in one offering. Printed at the top of
    every run so the registry gets maintained.
    """
    rows = fetch_all("""
        SELECT name, consecutive_failures, last_status,
               substr(coalesce(last_error, ''), 1, 70) AS err
          FROM trend_sources
         WHERE enabled AND consecutive_failures >= %s
         ORDER BY consecutive_failures DESC, name
    """, (UNHEALTHY_FAILURE_THRESHOLD,))

    if not rows:
        return []

    print()
    print(f"  WARNING — {len(rows)} source(s) failing repeatedly:")
    for r in rows:
        verdict = ("retire it" if r["consecutive_failures"] >= RETIRE_FAILURE_THRESHOLD
                   else "watch")
        print(f"    {_safe_console(r['name']):<26} {r['consecutive_failures']}x  "
              f"[{r['last_status']}] {_safe_console(r['err'])}  -> {verdict}")
    print("    Retire with: UPDATE trend_sources SET enabled=FALSE WHERE name='...';")
    return rows


def load_tracked_keywords() -> list[dict]:
    """
    Every keyword already in the tracked set, plus its token set.

    Used for novelty checking: proposing "salesforce consulting services"
    when we already track it wastes a reviewer's attention.
    """
    rows = fetch_all("""
        SELECT id, keyword, offering
          FROM keywords
         WHERE status = 'active'
    """)
    for r in rows:
        r["tokens"] = _token_set(r["keyword"])
    return rows


# ---------------------------------------------------------------------------
# Process — harvest
# ---------------------------------------------------------------------------

def harvest(
    sources: list[dict],
    lookback_days: int,
    max_items: int,
    dry_run: bool,
) -> tuple[list[dict], dict]:
    """
    Poll every source and persist new mentions.

    Returns (mentions, stats). Each mention dict carries its source metadata
    (name, category, weight) so scoring doesn't need a second join.

    Mentions already in the DB for a source are skipped — that's what makes
    re-running on the same day idempotent instead of doubling every count.
    """
    since = feeds.default_since(lookback_days)
    stats = {"sources_polled": 0, "sources_ok": 0, "sources_failed": 0,
             "items_fetched": 0, "mentions_new": 0, "mentions_duplicate": 0}

    collected: list[dict] = []

    for src, result in feeds.fetch_many(sources, max_items=max_items, since=since):
        stats["sources_polled"] += 1
        if result.status in ("error", "blocked"):
            stats["sources_failed"] += 1
            if not dry_run:
                _mark_source_failed(src["id"], result.status, result.error)
            continue

        stats["sources_ok"] += 1
        stats["items_fetched"] += len(result.items)

        for item in result.items:
            content_hash = _content_hash(item)
            record = {
                "source_id":       src["id"],
                "source_name":     src["name"],
                "source_category": src["category"],
                "source_weight":   float(src["weight"]),
                "offering_hint":   src.get("offering_hint"),
                "title":           item.title,
                "summary":         item.summary,
                "author":          item.author,
                "item_url":        item.url,
                "published_at":    item.published_at,
                "content_hash":    content_hash,
                "text":            item.text_blob(),
            }
            collected.append(record)

        if not dry_run:
            _mark_source_ok(src["id"], len(result.items))

    if dry_run:
        stats["mentions_new"] = len(collected)
        return collected, stats

    inserted_hashes = _insert_mentions(collected)
    stats["mentions_new"] = len(inserted_hashes)
    stats["mentions_duplicate"] = len(collected) - len(inserted_hashes)

    # Score the whole rolling window from the DB, not just this run's new
    # rows. Two reasons:
    #   - Stability. Re-running an hour later would otherwise see almost no
    #     new items and produce a near-empty candidate set, making the tool
    #     look broken when it is in fact working correctly.
    #   - Accumulation. A term mentioned once a day for two weeks is exactly
    #     the trend we want to catch, and no single run's delta reveals it.
    # Deduplication still happens at insert time, so a given article
    # contributes to the window exactly once no matter how often we poll.
    window = load_mentions_window(lookback_days, [s["id"] for s in sources])
    stats["mentions_in_window"] = len(window)
    return window, stats


def load_mentions_window(lookback_days: int, source_ids: list[int]) -> list[dict]:
    """
    Read the rolling window of harvested mentions, joined to source metadata.

    Falls back to `harvested_at` when a feed omitted a publication date —
    dropping those would quietly lose whole sources (Hacker News summaries
    and several vendor blogs routinely have no date).
    """
    return fetch_all("""
        SELECT m.source_id,
               m.title,
               m.summary,
               m.item_url,
               COALESCE(m.published_at, m.harvested_at) AS published_at,
               m.content_hash,
               s.name         AS source_name,
               s.category     AS source_category,
               s.weight       AS source_weight,
               s.offering_hint,
               concat_ws(' ', m.title, m.summary) AS text
          FROM trend_mentions m
          JOIN trend_sources s ON s.id = m.source_id
         WHERE m.source_id = ANY(%s)
           AND COALESCE(m.published_at, m.harvested_at) >= now() - make_interval(days => %s)
    """, (source_ids, lookback_days))


def _content_hash(item: feeds.FeedItem) -> str:
    basis = f"{_normalize(item.title)}|{(item.url or '').strip().lower()}"
    return hashlib.sha256(basis.encode("utf-8")).hexdigest()


def _insert_mentions(records: list[dict]) -> set[tuple[int, str]]:
    """
    Upsert mentions. Returns the (source_id, content_hash) pairs that were
    genuinely new.

    The `xmax = 0` trick distinguishes INSERT from UPDATE in a single
    round-trip: on a fresh insert xmax is 0; on a conflict-update it holds
    the updating transaction id.
    """
    if not records:
        return set()

    sql = """
        INSERT INTO trend_mentions
            (source_id, item_url, title, summary, author, published_at, content_hash, run_date)
        VALUES %s
        ON CONFLICT (source_id, content_hash) DO UPDATE
            SET title = EXCLUDED.title
        RETURNING source_id, content_hash, (xmax = 0) AS inserted
    """
    values = [
        (
            r["source_id"], r["item_url"], r["title"], r["summary"],
            r["author"], r["published_at"], r["content_hash"], date.today(),
        )
        for r in records
    ]

    new_pairs: set[tuple[int, str]] = set()
    with connection() as conn:
        with conn.cursor() as cur:
            rows = psycopg2.extras.execute_values(
                cur, sql, values, page_size=500, fetch=True,
            )
            for source_id, content_hash, inserted in rows:
                if inserted:
                    new_pairs.add((source_id, content_hash))
    return new_pairs


def _mark_source_ok(source_id: int, item_count: int) -> None:
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE trend_sources
                   SET last_polled_at       = now(),
                       last_status          = %s,
                       last_error           = NULL,
                       consecutive_failures = 0,
                       items_seen_total     = items_seen_total + %s,
                       updated_at           = now()
                 WHERE id = %s
            """, ("ok" if item_count else "empty", item_count, source_id))


def _mark_source_failed(source_id: int, status: str, error: str | None) -> None:
    with connection() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE trend_sources
                   SET last_polled_at       = now(),
                       last_status          = %s,
                       last_error           = %s,
                       consecutive_failures = consecutive_failures + 1,
                       updated_at           = now()
                 WHERE id = %s
            """, (status, (error or "")[:500], source_id))


# ---------------------------------------------------------------------------
# Process — phrase extraction
# ---------------------------------------------------------------------------

def _normalize(text: str | None) -> str:
    if not text:
        return ""
    return _WS_RE.sub(" ", text.lower()).strip()


def _tokenize(text: str) -> list[str]:
    return _TOKEN_RE.findall(_normalize(text))


def _token_set(text: str) -> frozenset[str]:
    return frozenset(t for t in _tokenize(text) if t not in STOPWORDS and len(t) > 2)


def extract_phrases(mentions: list[dict]) -> dict[str, dict]:
    """
    Pull candidate n-grams out of harvested text.

    Returns {normalized_phrase: aggregate}, where each aggregate tracks the
    mention count, which sources it appeared in, and up to 10 pieces of
    evidence (the articles it came from).

    A phrase survives extraction only if it:
      - is 2-5 tokens long
      - doesn't start or end with a stopword (so "of the cloud" is rejected
        while "cloud migration strategy" survives)
      - has at least one non-stopword token of real length
      - isn't pure boilerplate ("read more", "comments")

    Deliberately generous — the mention/spread thresholds and the offering
    classifier downstream do the real filtering. Better to over-extract here
    and filter with evidence than to guess at the extraction step.
    """
    agg: dict[str, dict] = {}

    for m in mentions:
        text = m.get("text") or ""
        if not text:
            continue
        # Sentence-ish segmentation so n-grams don't span punctuation and
        # invent phrases nobody wrote ("...the cloud. Microsoft announced..."
        # must not yield "cloud microsoft announced").
        # One credit per phrase per article — a phrase repeated six times in
        # one blog post is one outlet's emphasis, not six independent signals.
        # Scoped to the whole mention, not to each segment: a term in both the
        # title and the summary is still a single article, and counting it
        # twice inflates mention_count and duplicates the evidence trail.
        seen_in_mention: set[str] = set()

        for segment in re.split(r"[.!?;:|•\n\r()\[\]{}\"“”]+", text):
            tokens = _tokenize(segment)
            if len(tokens) < NGRAM_RANGE[0]:
                continue
            for n in range(NGRAM_RANGE[0], NGRAM_RANGE[1] + 1):
                for i in range(len(tokens) - n + 1):
                    gram = tokens[i:i + n]
                    if not _is_viable_phrase(gram):
                        continue
                    phrase = " ".join(gram)
                    if phrase in seen_in_mention:
                        continue
                    seen_in_mention.add(phrase)
                    _accumulate(agg, phrase, m)

    return agg


def _is_viable_phrase(gram: list[str]) -> bool:
    """
    Does this n-gram read like something a person would type into Google?

    Five gates, in the order that rejects the most for the least work.
    Applied during extraction, so the offering classifier and the paid
    volume lookup never see obvious prose.
    """
    # 1. Must not begin or end on a stopword: "of the cloud" out,
    #    "cloud migration strategy" in.
    if gram[0] in STOPWORDS or gram[-1] in STOPWORDS:
        return False

    # 2. Prose connectives anywhere mean it's a sentence fragment.
    if any(t in PROSE_MARKERS for t in gram):
        return False

    content = [t for t in gram if t not in STOPWORDS]
    if len(content) < 2:
        return False
    if all(len(t) <= 2 for t in content):
        return False

    # 3. Mostly digits — a date or a figure caption, not a keyword.
    if sum(1 for t in gram if t.isdigit()) > len(gram) / 2:
        return False

    # 4. Head-noun gate: reject "<our topic> + <news noun>" with nothing
    #    else to it. "ai models" and "cloud era" are headlines; "agentic ai"
    #    and "ai agents" survive because neither leans on a news noun.
    #    A commercial token ("services", "migration", "platform") always
    #    rescues the phrase — that's the shape of a real query.
    if not any(t in _commercial_tokens() for t in content):
        heads, markers = _generic_heads(), _offering_marker_words()
        if any(t in heads for t in content) and all(
            t in markers or t in heads for t in content
        ):
            return False

    return True


def _accumulate(agg: dict[str, dict], phrase: str, mention: dict) -> None:
    entry = agg.get(phrase)
    if entry is None:
        entry = {
            "phrase":         phrase,
            "mention_count":  0,
            "weighted_count": 0.0,
            "source_ids":     set(),
            "categories":     set(),
            "offering_hints": defaultdict(int),
            "first_seen":     None,
            "last_seen":      None,
            "evidence":       [],
        }
        agg[phrase] = entry

    entry["mention_count"] += 1
    # NUMERIC comes back from psycopg2 as Decimal; the in-memory harvest path
    # supplies a float. Normalize so arithmetic doesn't blow up on mixed types.
    entry["weighted_count"] += float(mention["source_weight"])
    entry["source_ids"].add(mention["source_id"])
    entry["categories"].add(mention["source_category"])
    if mention.get("offering_hint"):
        entry["offering_hints"][mention["offering_hint"]] += 1

    published = mention.get("published_at")
    if published:
        if entry["first_seen"] is None or published < entry["first_seen"]:
            entry["first_seen"] = published
        if entry["last_seen"] is None or published > entry["last_seen"]:
            entry["last_seen"] = published

    if len(entry["evidence"]) < 10:
        entry["evidence"].append({
            "title":        (mention.get("title") or "")[:200],
            "url":          mention.get("item_url"),
            "source":       mention.get("source_name"),
            "category":     mention.get("source_category"),
            "published_at": published.isoformat() if published else None,
        })


# ---------------------------------------------------------------------------
# Process — filtering and novelty
# ---------------------------------------------------------------------------

def filter_candidates(
    agg: dict[str, dict],
    tracked: list[dict],
    min_mentions: int,
    min_spread: int,
) -> list[dict]:
    """
    Cut the n-gram pile down to phrases worth proposing.

    Three gates, cheapest first:
      1. Frequency + source spread — a phrase must have been used by at
         least two different outlets.
      2. Commercial shape — either the phrase contains a service/intent
         token, or it maps cleanly to an offering. Pure news phrases
         ("announced a partnership") are dropped.
      3. Novelty — token-set similarity against the tracked keyword set.
    """
    tracked_exact = {_normalize(t["keyword"]) for t in tracked}
    survivors: list[dict] = []

    for phrase, entry in agg.items():
        # Long n-grams are usually one syndicated headline reproduced verbatim
        # across aggregators, not an independently recurring term. Require more
        # corroboration the longer the phrase gets.
        word_count = phrase.count(" ") + 1
        required_mentions = min_mentions + (1 if word_count >= 4 else 0)
        required_spread = min_spread + (1 if word_count >= 4 else 0)

        if entry["mention_count"] < required_mentions:
            continue
        if len(entry["source_ids"]) < required_spread:
            continue
        if phrase in tracked_exact:
            continue

        rule_offering = _classify_by_rule(phrase)
        has_commercial = bool(_token_set(phrase) & _commercial_tokens())
        if not rule_offering and not has_commercial:
            continue

        # A long phrase with no commercial token is a headline clause, not a
        # query — "release frontier ai models", "frontier ai model gemini".
        # Genuine long keywords earn their length with service words:
        # "application support and maintenance services".
        if word_count >= 4 and not has_commercial:
            continue

        nearest, similarity = _nearest_tracked(phrase, tracked)
        entry["rule_offering"] = rule_offering
        entry["has_commercial"] = has_commercial
        entry["nearest_tracked"] = nearest
        entry["nearest_similarity"] = similarity
        entry["is_novel"] = similarity < NOVELTY_SIMILARITY_THRESHOLD
        survivors.append(entry)

    # Rank by raw buzz so downstream caps (LLM, volume lookup) spend their
    # budget on the strongest signals.
    survivors.sort(key=lambda e: (-e["weighted_count"], -len(e["source_ids"])))
    return survivors


# A shorter phrase is dropped when a longer one containing it accounts for
# this share of its mentions — i.e. the short form was never really used on
# its own. 0.8 keeps "ai models" (16 mentions) alongside "open ai models" (3),
# while collapsing "open ai" (3) into "open ai models" (3).
SUBPHRASE_ABSORPTION_RATIO = 0.8


def collapse_subphrases(candidates: list[dict]) -> list[dict]:
    """
    Drop candidates that are contiguous sub-phrases of a longer candidate
    with comparable support.

    N-gram extraction over the same sentence necessarily emits every window,
    so one trending term yields a family of overlapping phrases ("open ai",
    "open ai models", "frontier ai", "frontier ai models"). Presenting all of
    them to a reviewer as separate keywords wastes attention and, worse,
    wastes paid volume lookups on near-identical queries.

    Keeps the more specific phrase when the shorter one adds no independent
    evidence. Expects `candidates` pre-sorted by descending weighted_count.
    """
    # Singular/plural pairs first: "ai coding tool" and "ai coding tools" are
    # one keyword to a reviewer, and two paid volume lookups to us. Keep the
    # better-supported form.
    by_stem: dict[tuple[str, ...], dict] = {}
    for c in candidates:
        stem = tuple(_singularize(t) for t in c["keyword"].split())
        prior = by_stem.get(stem)
        if prior is None or c["mention_count"] > prior["mention_count"]:
            by_stem[stem] = c
    candidates = list(by_stem.values())

    token_lists = [(c, c["keyword"].split()) for c in candidates]
    # Longest first, so a phrase is only ever measured against forms at least
    # as specific as itself.
    token_lists.sort(key=lambda pair: -len(pair[1]))

    kept: list[tuple[dict, list[str]]] = []
    for cand, tokens in token_lists:
        absorbed = False
        for longer, longer_tokens in kept:
            if len(longer_tokens) <= len(tokens):
                continue
            if not _is_contiguous_sublist(tokens, longer_tokens):
                continue
            if longer["mention_count"] >= SUBPHRASE_ABSORPTION_RATIO * cand["mention_count"]:
                absorbed = True
                break
        if not absorbed:
            kept.append((cand, tokens))

    return sorted((c for c, _ in kept), key=lambda c: -c["weighted_count"])


def _singularize(token: str) -> str:
    """
    Crude plural stripping, for grouping variants only.

    Never shown to a user and never written to the DB — it exists purely so
    "tools" and "tool" land in the same bucket. A real stemmer would be
    overkill and would mangle the technical vocabulary this pipeline deals
    in ("kubernetes", "devops", "analytics", "aws").
    """
    if len(token) <= 3 or not token.isalpha():
        return token
    if token.endswith("ss") or token.endswith("us") or token.endswith("is"):
        return token
    if token.endswith("ies") and len(token) > 4:
        return token[:-3] + "y"
    if token.endswith("es") and token[-3:-2] in ("s", "x", "z", "h"):
        return token[:-2]
    if token.endswith("s"):
        return token[:-1]
    return token


def _is_contiguous_sublist(needle: list[str], haystack: list[str]) -> bool:
    n = len(needle)
    return any(haystack[i:i + n] == needle for i in range(len(haystack) - n + 1))


def _classify_by_rule(phrase: str) -> str | None:
    """Longest matching offering token wins. Returns None if nothing matches."""
    padded = f" {phrase} "
    for token, offering in _offering_matchers():
        if f" {token} " in padded:
            return offering
    return None


# Phrases that open with these read as someone learning, not someone buying.
_INFORMATIONAL_OPENERS: tuple[str, ...] = (
    "what is", "what are", "how to", "how do", "how does", "why is", "why do",
    "when to", "guide to", "introduction to", "types of", "examples of",
    "difference between", "benefits of", "meaning of",
)

# Buying signals that go beyond a generic service noun.
_TRANSACTIONAL_TOKENS: frozenset[str] = frozenset(
    "pricing price cost quote hire buy purchase demo trial rates".split()
)


def _infer_intent(phrase: str, has_commercial: bool) -> str | None:
    """
    Rule-based intent, in the same vocabulary the LLM classifier uses
    (informational / commercial / transactional / navigational).

    Without this, only LLM-classified candidates ever carried an `intent`,
    so rule- and hint-classified ones could never earn the 1.15x commercial
    multiplier in score_candidate() — roughly 80% of candidates were
    systematically under-scored relative to whichever ones the LLM happened
    to handle. `intent` also lands in `keywords.intent` on promotion, so the
    gap outlived the trend run.
    """
    p = f" {phrase.strip().lower()} "
    tokens = set(p.split())

    if tokens & _TRANSACTIONAL_TOKENS:
        return "transactional"
    if any(p.startswith(f" {opener} ") for opener in _INFORMATIONAL_OPENERS):
        return "informational"
    if has_commercial or (tokens & _commercial_tokens()):
        return "commercial"
    return None


def _nearest_tracked(phrase: str, tracked: list[dict]) -> tuple[str | None, float]:
    """
    Closest tracked keyword by token-set Jaccard similarity.

    Jaccard rather than edit distance because keyword duplication is about
    shared concepts, not shared characters: "salesforce implementation
    services" and "services for salesforce implementation" are the same
    keyword, and their edit distance is large.

    Blind to paraphrase, though — see `_nearest_tracked_semantic`, which
    supersedes this when embeddings are available.
    """
    phrase_tokens = _token_set(phrase)
    if not phrase_tokens:
        return None, 0.0

    best_kw: str | None = None
    best_sim = 0.0
    for t in tracked:
        other = t["tokens"]
        if not other:
            continue
        union = len(phrase_tokens | other)
        if not union:
            continue
        sim = len(phrase_tokens & other) / union
        if sim > best_sim:
            best_sim, best_kw = sim, t["keyword"]
    return best_kw, round(best_sim, 3)


def build_semantic_index(phrases: list[str], tracked: list[dict]) -> dict | None:
    """
    Embed the candidate phrases and the tracked keyword set once per run.

    Returns a lookup for `_nearest_tracked_semantic`, or None when embeddings
    are unavailable — in which case callers keep using Jaccard and the run
    proceeds exactly as before.

    Batched deliberately: one call covering everything, then pure arithmetic
    per comparison. Embedding per candidate inside the scoring loop would turn
    a cent into a bill.
    """
    if not phrases or not tracked:
        return None

    from common.connectors.embeddings import embed_cached, is_available

    if not is_available():
        logger.debug("Embeddings unavailable — novelty falls back to Jaccard")
        return None

    tracked_keywords = [t["keyword"] for t in tracked if t.get("keyword")]
    vectors = embed_cached(tracked_keywords + list(phrases))
    if not vectors:
        return None

    tracked_vectors = {k: v for k, v in vectors.items() if k in set(tracked_keywords)}
    if not tracked_vectors:
        return None

    return {"tracked": tracked_vectors, "all": vectors}


def refine_novelty_semantic(candidates: list[dict], tracked: list[dict]) -> dict:
    """
    Re-check novelty for the surviving candidates using embeddings.

    Runs once, over the final candidate list, after Jaccard has already done
    the cheap pass. Only ever makes a candidate *less* novel — the token
    overlap Jaccard found is real, so a high Jaccard score is never overridden
    downward by a lower cosine.

    Returns counters. Silently a no-op when embeddings are unavailable.
    """
    stats = {"checked": 0, "demoted": 0, "reason": None}
    if not candidates or not tracked:
        stats["reason"] = "nothing to compare"
        return stats

    index = build_semantic_index([c["keyword"] for c in candidates], tracked)
    if index is None:
        stats["reason"] = "embeddings unavailable — Jaccard result kept"
        return stats

    for c in candidates:
        got = _nearest_tracked_semantic(c["keyword"], index)
        if got is None:
            continue
        label, score = got
        stats["checked"] += 1

        # Keep whichever signal says "less novel". Jaccard finding real token
        # overlap is evidence in its own right; embeddings add paraphrase
        # detection on top rather than replacing it.
        if score <= (c.get("nearest_similarity") or 0.0):
            continue

        was_novel = c.get("is_novel", True)
        c["nearest_tracked"] = label
        c["nearest_similarity"] = score
        c["is_novel"] = score < NOVELTY_SIMILARITY_THRESHOLD
        if was_novel and not c["is_novel"]:
            stats["demoted"] += 1
            logger.info("Semantic near-duplicate: %r ~ tracked %r (%.2f)",
                        c["keyword"], label, score)
    return stats


def _nearest_tracked_semantic(phrase: str, index: dict) -> tuple[str | None, float] | None:
    """
    Closest tracked keyword by cosine similarity over embeddings.

    Returns None when this particular phrase has no vector — the caller then
    falls back to Jaccard for that one phrase rather than the whole run.

    This is the guard on recurring spend: a promoted keyword costs money on
    every future rank-tracker run, and Jaccard was waving through paraphrases
    of keywords already tracked.
    """
    from common.connectors.embeddings import nearest

    vector = index["all"].get(phrase.strip())
    if vector is None:
        return None
    label, score = nearest(vector, index["tracked"])
    return label, round(score, 3)


# ---------------------------------------------------------------------------
# Process — offering classification (rule first, Claude for the residue)
# ---------------------------------------------------------------------------

def _classifier_system() -> str:
    """
    The system prompt. Brand identity enters here and nowhere else — the user
    prompt below stays tenant-neutral by construction.
    """
    return system_preamble(
        "You are an SEO keyword strategist. You classify emerging industry "
        "phrases into service offerings and reshape them into realistic search "
        "queries that a buyer would actually type."
    )


def build_classifier_prompt(phrases: list[dict], offerings: list[str]) -> str:
    """
    Ask Claude to do the one thing rules can't: judge what an unfamiliar
    phrase is about, and reshape a news-shaped phrase into a search-shaped one.
    """
    lines = [
        "Below are phrases trending in industry publications and practitioner "
        "communities. These are the service offerings in scope:",
        "",
        ", ".join(offerings),
        "",
        "For each phrase, decide:",
        "",
        "1. `offering` — which offering above it belongs to, or \"none\" if it is "
        "general tech news with no connection to any of them (M&A, funding rounds, "
        "layoffs, executive moves, product launches by unrelated vendors, "
        "consumer tech, politics).",
        "2. `keyword` — the phrase rewritten as a search query a B2B buyer would "
        "type when looking to hire a services partner. Keep it 2-6 words. Keep the "
        "distinctive technical term from the original. Add a commercial modifier "
        "(services, solutions, consulting, platform, software, company) only when "
        "the original lacks one. Do NOT invent a topic that isn't in the phrase.",
        "3. `intent` — one of: informational, commercial, transactional, navigational.",
        "4. `confidence` — 0.0 to 1.0, how sure you are about the offering.",
        "",
        "Be strict with \"none\". A phrase must describe technology or a business "
        "problem one of the offerings above could be sold against. When in doubt, "
        "answer \"none\" — a false positive wastes an analyst's review time and a "
        "rank-tracking budget.",
        "",
        "Phrases:",
    ]
    for i, entry in enumerate(phrases, 1):
        sample = entry["evidence"][0]["title"] if entry["evidence"] else ""
        lines.append(
            f'{i}. "{entry["phrase"]}"  '
            f'(seen {entry["mention_count"]}x across {len(entry["source_ids"])} sources; '
            f'e.g. "{sample[:110]}")'
        )

    lines += [
        "",
        "Return ONLY a JSON array, no prose, no markdown fence. One object per "
        "phrase, in the same order:",
        "",
        '[{"n": 1, "offering": "AI", "keyword": "agentic ai development services", '
        '"intent": "commercial", "confidence": 0.9}, ...]',
    ]
    return "\n".join(lines)


def classify_with_llm(phrases: list[dict], offerings: list[str]) -> dict[str, dict]:
    """
    Classify phrases the token rules couldn't place.

    Returns {phrase: {offering, keyword, intent, confidence}}. On any LLM
    failure returns {} — the caller keeps its rule-based results and the run
    continues. This is the graceful-degradation contract in common/llm.py.
    """
    if not phrases:
        return {}

    prompt = build_classifier_prompt(phrases, offerings)
    try:
        text, usage = call_claude(
            prompt,
            tier="cheap",           # classification is exactly what haiku is for
            system=_classifier_system(),
            max_tokens=8000,
            temperature=0.2,        # low: we want consistent labels, not creativity
        )
    except LLMUnavailableError as exc:
        logger.warning("LLM classification unavailable (%s) — keeping rule-based labels", exc)
        return {}

    logger.info("LLM classification: %d phrase(s), est cost $%.4f",
                len(phrases), usage.get("est_cost_usd", 0.0))

    parsed = _parse_json_array(text)
    if parsed is None:
        logger.warning("LLM returned unparseable output — keeping rule-based labels")
        return {}

    out: dict[str, dict] = {}
    for obj in parsed:
        try:
            idx = int(obj.get("n", 0)) - 1
        except (TypeError, ValueError):
            continue
        if not (0 <= idx < len(phrases)):
            continue
        offering = (obj.get("offering") or "").strip()
        if not offering or offering.lower() == "none":
            continue
        if offering not in offerings:
            logger.debug("LLM proposed unknown offering %r — ignoring", offering)
            continue
        keyword = _normalize(obj.get("keyword") or phrases[idx]["phrase"])
        if not keyword:
            continue
        out[phrases[idx]["phrase"]] = {
            "offering":   offering,
            "keyword":    keyword,
            "intent":     (obj.get("intent") or "").strip().lower() or None,
            "confidence": _clamp_float(obj.get("confidence"), 0.0, 1.0, default=0.6),
        }
    return out


def _parse_json_array(text: str) -> list[dict] | None:
    """Extract a JSON array from a model response, fenced or not."""
    text = (text or "").strip()
    fence = re.search(r"```(?:json)?\s*(.+?)```", text, re.DOTALL)
    if fence:
        text = fence.group(1).strip()
    start, end = text.find("["), text.rfind("]")
    if start == -1 or end == -1 or end < start:
        return None
    try:
        parsed = json.loads(text[start:end + 1])
    except json.JSONDecodeError:
        return None
    return parsed if isinstance(parsed, list) else None


def _clamp_float(value, low: float, high: float, default: float) -> float:
    try:
        return max(low, min(high, float(value)))
    except (TypeError, ValueError):
        return default


# ---------------------------------------------------------------------------
# Process — search volume (Google Ads Keyword Planner via DataForSEO)
# ---------------------------------------------------------------------------

def enrich_with_volume(candidates: list[dict], max_lookups: int) -> dict:
    """
    Attach Keyword Planner volume, CPC, competition, and 12-month trend.

    Batches at the API cap so a 600-keyword lookup is one $0.05 task, not 600.
    Failure is non-fatal: candidates keep their buzz scores and get
    volume_source=None, which the scorer handles.
    """
    lookups = [c for c in candidates[:max_lookups]]
    if not lookups:
        return {"looked_up": 0, "matched": 0, "cost_usd": 0.0, "error": None}

    terms = [c["keyword"] for c in lookups]
    p = profile()
    try:
        # Explicit market: volume for the tenant's location/language, not
        # whatever the connector happens to default to.
        results = get_search_volume(
            terms, location_code=p.location_code, language_code=p.language_code
        )
    except DataForSEOError as exc:
        logger.warning("Keyword Planner lookup failed (non-fatal): %s", exc)
        return {"looked_up": len(terms), "matched": 0, "cost_usd": 0.0, "error": str(exc)}

    by_keyword = {_normalize(r["keyword"]): r for r in results if r.get("keyword")}
    matched = 0
    for c in lookups:
        data = by_keyword.get(_normalize(c["keyword"]))
        if not data:
            continue
        matched += 1
        c["search_volume"]     = data.get("search_volume")
        c["cpc"]               = data.get("cpc")
        c["competition"]       = data.get("competition")
        c["competition_index"] = data.get("competition_index")
        c["monthly_searches"]  = data.get("monthly_searches") or []
        c["momentum_ratio"]    = compute_momentum(c["monthly_searches"])
        c["volume_source"]     = data.get("source")
        c["volume_checked_at"] = datetime.now(timezone.utc)

    batches = math.ceil(len(terms) / KEYWORD_PLANNER_BATCH_SIZE)
    return {
        "looked_up": len(terms),
        "matched":   matched,
        "cost_usd":  round(batches * COST_PER_KEYWORD_PLANNER_TASK, 4),
        "error":     None,
    }


def compute_momentum(monthly: list[dict]) -> float | None:
    """
    Mean volume of the last 3 months / mean of the prior 9.

    >1.0 means demand is rising. This is the number that separates "a big
    keyword" from "a keyword that is becoming big" — the whole point of a
    trend scout. Needs a full 12-month array to be meaningful.
    """
    if not monthly or len(monthly) < 12:
        return None
    volumes = [m.get("search_volume") or 0 for m in monthly]
    recent, prior = volumes[-3:], volumes[-12:-3]
    prior_mean = sum(prior) / len(prior) if prior else 0
    if prior_mean <= 0:
        # No prior demand at all but present demand now: genuinely emergent.
        return 3.0 if sum(recent) > 0 else None
    recent_mean = sum(recent) / len(recent)
    return round(min(recent_mean / prior_mean, 10.0), 3)


# ---------------------------------------------------------------------------
# Process — scoring
# ---------------------------------------------------------------------------

def score_candidate(c: dict) -> dict:
    """
    Five 0-100 sub-scores, combined by SCORE_WEIGHTS into trend_score.

    Each sub-score answers one question:
      buzz        Is the industry talking about it, in more than one place?
      volume      Does anyone actually search it?
      momentum    Is that search demand growing?
      opportunity Is this new territory, or a rewording of what we track?
      commercial  Would the traffic be worth anything?

    Missing volume data doesn't zero a candidate out — it scores the neutral
    midpoint, so a strong buzz signal still surfaces for review when the
    Planner has no data (which is common for genuinely new terms).
    """
    # --- buzz: log-damped mention count, scaled by source and category spread
    weighted = c.get("weighted_count", 0.0)
    source_spread = c.get("source_spread", 0)
    category_spread = c.get("category_spread", 1)
    buzz_base = min(100.0, 28.0 * math.log1p(weighted))
    spread_mult = CATEGORY_SPREAD_BONUS.get(min(category_spread, 4), 1.0)
    buzz = min(100.0, buzz_base * spread_mult * (1 + 0.06 * max(0, source_spread - 2)))

    # --- volume: log scale. 10/mo ≈ 25, 100 ≈ 50, 1k ≈ 75, 10k+ ≈ 100.
    sv = c.get("search_volume")
    if sv is None:
        volume = 40.0          # unknown, not zero — see docstring
    elif sv <= 0:
        volume = 0.0
    else:
        volume = min(100.0, 25.0 * math.log10(sv * 10))

    # --- momentum: 1.0 (flat) = 40, 2.0 (doubled) = 80, 3.0+ = 100
    ratio = c.get("momentum_ratio")
    if ratio is None:
        momentum = 40.0
    else:
        momentum = min(100.0, max(0.0, 40.0 * ratio))

    # --- opportunity: novelty vs the tracked set
    similarity = float(c.get("nearest_similarity") or 0.0)
    opportunity = max(0.0, 100.0 * (1.0 - similarity))
    if not c.get("is_novel", True):
        opportunity *= 0.3     # near-duplicate: heavily penalized, not excluded

    # --- commercial: CPC is the market's own estimate of a click's worth
    cpc = c.get("cpc")
    if cpc is None:
        commercial = 45.0 if c.get("has_commercial") else 30.0
    else:
        commercial = min(100.0, 20.0 * math.log1p(float(cpc)) * 2)
    if c.get("intent") in ("commercial", "transactional"):
        commercial = min(100.0, commercial * 1.15)

    trend = (
        buzz        * SCORE_WEIGHTS["buzz"]
        + volume    * SCORE_WEIGHTS["volume"]
        + momentum  * SCORE_WEIGHTS["momentum"]
        + opportunity * SCORE_WEIGHTS["opportunity"]
        + commercial  * SCORE_WEIGHTS["commercial"]
    ) / 100.0

    c["buzz_score"]        = round(buzz, 2)
    c["volume_score"]      = round(volume, 2)
    c["momentum_score"]    = round(momentum, 2)
    c["opportunity_score"] = round(opportunity, 2)
    c["commercial_score"]  = round(commercial, 2)
    c["trend_score"]       = round(trend, 2)
    return c


# ---------------------------------------------------------------------------
# Write
# ---------------------------------------------------------------------------

def upsert_candidates(candidates: list[dict]) -> dict:
    """
    Persist candidates. Returns {"inserted": n, "updated": n}.

    Re-running accumulates evidence rather than replacing it: mention counts
    take the larger value and first_seen keeps the earlier date, so a
    candidate that keeps reappearing across weekly runs shows a growing
    footprint. A human's `status` and `review_note` are never overwritten —
    rejecting a candidate makes it stay rejected.
    """
    if not candidates:
        return {"inserted": 0, "updated": 0}

    sql = """
        INSERT INTO keyword_candidates (
            candidate_keyword, display_keyword, source_phrase,
            suggested_offering, offering_confidence, classification_method, intent,
            mention_count, source_spread, category_spread, first_seen_at, last_seen_at,
            search_volume, cpc, competition, competition_index,
            monthly_searches, momentum_ratio, volume_source, volume_checked_at,
            is_novel, nearest_tracked_keyword, nearest_similarity,
            buzz_score, volume_score, momentum_score, opportunity_score,
            commercial_score, trend_score,
            evidence, last_scored_date
        ) VALUES %s
        ON CONFLICT (candidate_keyword) DO UPDATE SET
            display_keyword         = EXCLUDED.display_keyword,
            suggested_offering      = COALESCE(EXCLUDED.suggested_offering,
                                               keyword_candidates.suggested_offering),
            offering_confidence     = EXCLUDED.offering_confidence,
            classification_method   = EXCLUDED.classification_method,
            intent                  = COALESCE(EXCLUDED.intent, keyword_candidates.intent),
            mention_count           = GREATEST(EXCLUDED.mention_count,
                                               keyword_candidates.mention_count),
            source_spread           = GREATEST(EXCLUDED.source_spread,
                                               keyword_candidates.source_spread),
            category_spread         = GREATEST(EXCLUDED.category_spread,
                                               keyword_candidates.category_spread),
            first_seen_at           = LEAST(EXCLUDED.first_seen_at,
                                            keyword_candidates.first_seen_at),
            last_seen_at            = GREATEST(EXCLUDED.last_seen_at,
                                               keyword_candidates.last_seen_at),
            -- Volume fields: keep the previous value when this run skipped
            -- the paid lookup (--no-volume), rather than nulling good data.
            search_volume           = COALESCE(EXCLUDED.search_volume,
                                               keyword_candidates.search_volume),
            cpc                     = COALESCE(EXCLUDED.cpc, keyword_candidates.cpc),
            competition             = COALESCE(EXCLUDED.competition,
                                               keyword_candidates.competition),
            competition_index       = COALESCE(EXCLUDED.competition_index,
                                               keyword_candidates.competition_index),
            monthly_searches        = COALESCE(EXCLUDED.monthly_searches,
                                               keyword_candidates.monthly_searches),
            momentum_ratio          = COALESCE(EXCLUDED.momentum_ratio,
                                               keyword_candidates.momentum_ratio),
            volume_source           = COALESCE(EXCLUDED.volume_source,
                                               keyword_candidates.volume_source),
            volume_checked_at       = COALESCE(EXCLUDED.volume_checked_at,
                                               keyword_candidates.volume_checked_at),
            is_novel                = EXCLUDED.is_novel,
            nearest_tracked_keyword = EXCLUDED.nearest_tracked_keyword,
            nearest_similarity      = EXCLUDED.nearest_similarity,
            buzz_score              = EXCLUDED.buzz_score,
            volume_score            = EXCLUDED.volume_score,
            momentum_score          = EXCLUDED.momentum_score,
            opportunity_score       = EXCLUDED.opportunity_score,
            commercial_score        = EXCLUDED.commercial_score,
            trend_score             = EXCLUDED.trend_score,
            evidence                = EXCLUDED.evidence,
            last_scored_date        = EXCLUDED.last_scored_date,
            updated_at              = now()
            -- status / reviewed_by / review_note deliberately untouched:
            -- a human decision outlives any number of re-scoring runs.
        RETURNING (xmax = 0) AS inserted
    """

    values = []
    for c in candidates:
        values.append((
            c["keyword"],
            c.get("display_keyword") or c["keyword"],
            c.get("phrase"),
            c.get("offering"),
            c.get("offering_confidence"),
            c.get("classification_method"),
            c.get("intent"),
            c.get("mention_count", 0),
            c.get("source_spread", 0),
            c.get("category_spread", 0),
            c.get("first_seen"),
            c.get("last_seen"),
            c.get("search_volume"),
            c.get("cpc"),
            c.get("competition"),
            c.get("competition_index"),
            psycopg2.extras.Json(c["monthly_searches"]) if c.get("monthly_searches") else None,
            c.get("momentum_ratio"),
            c.get("volume_source"),
            c.get("volume_checked_at"),
            c.get("is_novel", True),
            c.get("nearest_tracked"),
            c.get("nearest_similarity"),
            c.get("buzz_score", 0),
            c.get("volume_score", 0),
            c.get("momentum_score", 0),
            c.get("opportunity_score", 0),
            c.get("commercial_score", 0),
            c.get("trend_score", 0),
            psycopg2.extras.Json(c.get("evidence") or []),
            date.today(),
        ))

    inserted = 0
    with connection() as conn:
        with conn.cursor() as cur:
            rows = psycopg2.extras.execute_values(
                cur, sql, values, page_size=200, fetch=True,
            )
            inserted = sum(1 for (was_insert,) in rows if was_insert)

    return {"inserted": inserted, "updated": len(values) - inserted}


# ---------------------------------------------------------------------------
# Human gate — promotion into the tracked keyword set
# ---------------------------------------------------------------------------

def promote_candidates(
    candidate_ids: list[int] | None = None,
    min_score: float | None = None,
    offering: str | None = None,
    dry_run: bool = False,
) -> dict:
    """
    Move candidates into `keywords` so the rank tracker picks them up.

    This is the only path from discovery into the tracked set, and it only
    runs when a human invokes `--promote`. Explicit IDs are the intended
    usage; `--min-score` exists for bulk promotion after a review pass and
    requires `status='approved'` so it can't sweep up unreviewed noise.

    Never deletes. Never rewrites an existing keyword's offering. A candidate
    whose keyword already exists in `keywords` is marked 'duplicate' and
    linked to the existing row.
    """
    if not candidate_ids and min_score is None:
        raise ValueError("promote_candidates needs either candidate_ids or min_score")

    sql = "SELECT * FROM keyword_candidates WHERE status <> 'promoted'"
    params: list = []
    if candidate_ids:
        sql += " AND id = ANY(%s)"
        params.append(candidate_ids)
    else:
        # Bulk path: only approved candidates. Score alone is not consent.
        sql += " AND status = 'approved' AND trend_score >= %s"
        params.append(min_score)
    if offering:
        sql += " AND suggested_offering = %s"
        params.append(offering)
    sql += " ORDER BY trend_score DESC"

    rows = fetch_all(sql, params)
    if not rows:
        return {"promoted": 0, "duplicates": 0, "skipped": 0, "details": []}

    promoted = duplicates = skipped = 0
    details: list[dict] = []

    for row in rows:
        keyword = row["candidate_keyword"]
        offering_val = row["suggested_offering"]
        if not offering_val:
            skipped += 1
            details.append({"id": row["id"], "keyword": keyword,
                            "action": "skipped", "reason": "no suggested_offering"})
            continue

        if dry_run:
            promoted += 1
            details.append({"id": row["id"], "keyword": keyword,
                            "action": "would promote", "offering": offering_val})
            continue

        with connection() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
                # keywords has UNIQUE(keyword, offering). DO NOTHING then
                # SELECT tells us whether we created it or it already existed.
                cur.execute("""
                    INSERT INTO keywords (keyword, offering, intent, type, status)
                    VALUES (%s, %s, %s, 'primary', 'active')
                    ON CONFLICT (keyword, offering) DO NOTHING
                    RETURNING id
                """, (keyword, offering_val, row["intent"]))
                created = cur.fetchone()

                if created:
                    keyword_id = created["id"]
                    new_status = "promoted"
                    promoted += 1
                    action = "promoted"
                else:
                    cur.execute(
                        "SELECT id FROM keywords WHERE keyword = %s AND offering = %s",
                        (keyword, offering_val),
                    )
                    existing = cur.fetchone()
                    keyword_id = existing["id"] if existing else None
                    new_status = "duplicate"
                    duplicates += 1
                    action = "already tracked"

                cur.execute("""
                    UPDATE keyword_candidates
                       SET status              = %s,
                           promoted_keyword_id = %s,
                           reviewed_at         = COALESCE(reviewed_at, now()),
                           updated_at          = now()
                     WHERE id = %s
                """, (new_status, keyword_id, row["id"]))

        details.append({"id": row["id"], "keyword": keyword,
                        "action": action, "offering": offering_val,
                        "keyword_id": keyword_id})

    return {"promoted": promoted, "duplicates": duplicates,
            "skipped": skipped, "details": details}


def list_candidates(min_score: float = 0.0, offering: str | None = None,
                    status: str = "new", limit: int = 50) -> list[dict]:
    sql = """
        SELECT id, display_keyword, suggested_offering, search_volume, cpc,
               competition, momentum_ratio, mention_count, source_spread,
               trend_score, status, is_novel, nearest_tracked_keyword,
               first_discovered_date
          FROM keyword_candidates
         WHERE trend_score >= %s
    """
    params: list = [min_score]
    if status and status != "all":
        sql += " AND status = %s"
        params.append(status)
    if offering:
        sql += " AND suggested_offering = %s"
        params.append(offering)
    sql += " ORDER BY trend_score DESC LIMIT %s"
    params.append(limit)
    return fetch_all(sql, params)


# ---------------------------------------------------------------------------
# Notify
# ---------------------------------------------------------------------------

def print_candidate_table(rows: list[dict], title: str) -> None:
    print()
    print("=" * 118)
    print(title)
    print("=" * 118)
    if not rows:
        print("  (none)")
        return
    print(f"  {'id':>5}  {'score':>6}  {'vol':>7}  {'mom':>5}  {'ment':>5}  "
          f"{'offering':<18}  keyword")
    print(f"  {'-'*5}  {'-'*6}  {'-'*7}  {'-'*5}  {'-'*5}  {'-'*18}  {'-'*40}")
    for r in rows:
        sv = r.get("search_volume")
        mom = r.get("momentum_ratio")
        print(
            f"  {r['id']:>5}  {float(r['trend_score']):>6.1f}  "
            f"{(f'{sv:,}' if sv is not None else '—'):>7}  "
            f"{(f'{float(mom):.2f}' if mom is not None else '—'):>5}  "
            f"{r.get('mention_count', 0):>5}  "
            f"{_safe_console(str(r.get('suggested_offering') or '?'))[:18]:<18}  "
            f"{_safe_console(r.get('display_keyword'))[:52]}"
        )


def write_reports(candidates: list[dict], run_date: date, stats: dict) -> dict[str, Path]:
    """Excel workbook + markdown digest. Returns {kind: path}."""
    out_dir = settings.OUTPUTS_DIR / "reports"
    out_dir.mkdir(parents=True, exist_ok=True)
    written: dict[str, Path] = {}

    ranked = sorted(candidates, key=lambda c: -c.get("trend_score", 0))

    xlsx_path = out_dir / f"trend_scout_{run_date.isoformat()}.xlsx"
    try:
        _write_excel(ranked, xlsx_path, run_date, stats)
        written["excel"] = xlsx_path
    except Exception as exc:                       # openpyxl missing, file locked
        logger.warning("Excel report failed (non-fatal): %s", exc)

    md_path = settings.OUTPUTS_DIR / "audits" / f"trend_digest_{run_date.isoformat()}.md"
    try:
        md_path.parent.mkdir(parents=True, exist_ok=True)
        md_path.write_text(_build_digest(ranked, run_date, stats), encoding="utf-8")
        written["markdown"] = md_path
    except OSError as exc:
        logger.warning("Markdown digest failed (non-fatal): %s", exc)

    return written


def _write_excel(ranked: list[dict], path: Path, run_date: date, stats: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1B3A5C", end_color="1B3A5C")
    hot_fill    = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
    warm_fill   = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")

    headers = [
        "ID", "Candidate Keyword", "Suggested Offering", "Intent",
        "Trend Score", "Search Volume", "Momentum", "CPC", "Competition",
        "Mentions", "Sources", "Categories",
        "Buzz", "Volume Score", "Momentum Score", "Opportunity", "Commercial",
        "Novel?", "Nearest Tracked Keyword", "Similarity",
        "Classified By", "Top Evidence",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, c in enumerate(ranked, 2):
        score = c.get("trend_score", 0)
        evidence = c.get("evidence") or []
        top = evidence[0] if evidence else {}
        values = [
            c.get("db_id"), c.get("display_keyword") or c.get("keyword"),
            c.get("offering"), c.get("intent"),
            score, c.get("search_volume"), c.get("momentum_ratio"),
            float(c["cpc"]) if c.get("cpc") is not None else None,
            c.get("competition"),
            c.get("mention_count"), c.get("source_spread"), c.get("category_spread"),
            c.get("buzz_score"), c.get("volume_score"), c.get("momentum_score"),
            c.get("opportunity_score"), c.get("commercial_score"),
            "Yes" if c.get("is_novel", True) else "No",
            c.get("nearest_tracked"), c.get("nearest_similarity"),
            c.get("classification_method"),
            (top.get("title") or "")[:180],
        ]
        for col, v in enumerate(values, 1):
            ws.cell(row=i, column=col, value=v)
        if score >= 60:
            ws.cell(row=i, column=5).fill = hot_fill
        elif score >= 45:
            ws.cell(row=i, column=5).fill = warm_fill

    widths = [7, 42, 18, 14, 11, 13, 10, 9, 12, 9, 9, 11,
              8, 12, 14, 12, 11, 8, 40, 11, 13, 60]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions

    # Sheet 2 — per-offering rollup
    ws2 = wb.create_sheet("By Offering")
    by_off: dict[str, list[dict]] = defaultdict(list)
    for c in ranked:
        by_off[c.get("offering") or "(unclassified)"].append(c)

    for col, h in enumerate(["Offering", "Candidates", "Avg Score", "Top Candidate",
                             "Top Score", "Total Search Volume"], 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for i, (off, items) in enumerate(
        sorted(by_off.items(), key=lambda kv: -len(kv[1])), 2
    ):
        scores = [c.get("trend_score", 0) for c in items]
        ws2.cell(row=i, column=1, value=off)
        ws2.cell(row=i, column=2, value=len(items))
        ws2.cell(row=i, column=3, value=round(sum(scores) / len(scores), 1) if scores else 0)
        ws2.cell(row=i, column=4, value=items[0].get("display_keyword") or items[0].get("keyword"))
        ws2.cell(row=i, column=5, value=items[0].get("trend_score"))
        ws2.cell(row=i, column=6, value=sum(c.get("search_volume") or 0 for c in items))
    for col, w in [(1, 22), (2, 12), (3, 11), (4, 44), (5, 11), (6, 20)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    # Sheet 3 — run provenance, so a reader can tell how the numbers were made
    ws3 = wb.create_sheet("Run Info")
    ws3.cell(row=1, column=1, value=f"Trend Scout — {run_date.isoformat()}").font = \
        Font(name="Arial", bold=True, size=14)
    for i, (k, v) in enumerate(stats.items(), 3):
        ws3.cell(row=i, column=1, value=k.replace("_", " ").title())
        ws3.cell(row=i, column=2, value=str(v))
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 60

    wb.save(path)


def _build_digest(ranked: list[dict], run_date: date, stats: dict) -> str:
    lines = [
        f"# Trend Scout digest — {run_date.isoformat()}",
        "",
        f"_Harvested {stats.get('items_fetched', 0)} items from "
        f"{stats.get('sources_ok', 0)}/{stats.get('sources_polled', 0)} sources "
        f"over the last {stats.get('lookback_days', DEFAULT_LOOKBACK_DAYS)} days._",
        "",
        "## Top emerging keywords",
        "",
        "| Score | Keyword | Offering | Volume | Momentum | Mentions |",
        "|---:|---|---|---:|---:|---:|",
    ]
    for c in ranked[:30]:
        sv = c.get("search_volume")
        mom = c.get("momentum_ratio")
        lines.append(
            f"| {c.get('trend_score', 0):.1f} "
            f"| {c.get('display_keyword') or c.get('keyword')} "
            f"| {c.get('offering') or '—'} "
            f"| {f'{sv:,}' if sv is not None else '—'} "
            f"| {f'{float(mom):.2f}x' if mom is not None else '—'} "
            f"| {c.get('mention_count', 0)} |"
        )

    by_off: dict[str, list[dict]] = defaultdict(list)
    for c in ranked:
        by_off[c.get("offering") or "(unclassified)"].append(c)

    lines += ["", "## By offering", "",
              "| Offering | Candidates | Best candidate | Score |",
              "|---|---:|---|---:|"]
    for off, items in sorted(by_off.items(), key=lambda kv: -len(kv[1])):
        best = items[0]
        lines.append(
            f"| {off} | {len(items)} "
            f"| {best.get('display_keyword') or best.get('keyword')} "
            f"| {best.get('trend_score', 0):.1f} |"
        )

    lines += [
        "",
        "## How to act on this",
        "",
        "1. Review the queue: `SELECT * FROM v_trend_review_queue LIMIT 40;`",
        "2. Approve the ones worth tracking:",
        "   `UPDATE keyword_candidates SET status='approved', reviewed_by='<you>', "
        "reviewed_at=now() WHERE id IN (...);`",
        "3. Promote them into the tracked set:",
        "   `python -m keyword_intelligence.trend_scout --promote --ids <id,id,id>`",
        "4. Rank-track the newly added keywords:",
        "   `python -m keyword_intelligence.rank_tracker --offering \"<offering>\" --all`",
        "",
        "Promotion is deliberately manual — every tracked keyword adds recurring "
        "DataForSEO cost to every future rank-tracker run.",
        "",
        "---",
        "",
        "## Scoring model",
        "",
        "| Component | Weight | What it measures |",
        "|---|---:|---|",
        "| Buzz | 30 | Mentions, weighted by source trust and spread across outlets |",
        "| Volume | 25 | Google Ads Keyword Planner monthly searches (log-scaled) |",
        "| Momentum | 20 | Last 3 months vs prior 9 — is demand rising? |",
        "| Opportunity | 15 | Distance from the keywords we already track |",
        "| Commercial | 10 | CPC and intent — would the traffic be worth anything? |",
        "",
        f"_Generated by `{AGENT_NAME}`._",
    ]
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def run(
    lookback_days: int = DEFAULT_LOOKBACK_DAYS,
    offering: str | None = None,
    max_items: int = DEFAULT_MAX_ITEMS_PER_SOURCE,
    min_mentions: int = MIN_MENTIONS,
    min_spread: int = MIN_SOURCE_SPREAD,
    max_volume_lookups: int = DEFAULT_MAX_VOLUME_LOOKUPS,
    use_llm: bool = True,
    use_volume: bool = True,
    dry_run: bool = False,
    skip_reports: bool = False,
) -> dict:
    started = time.monotonic()
    run_date = date.today()
    errors: list[str] = []

    # ---- Read -------------------------------------------------------------
    sources = load_sources(offering=offering)
    if not sources:
        msg = (f"No enabled sources"
               + (f" with offering_hint={offering!r}" if offering else "")
               + ". Check `SELECT * FROM trend_sources;` — has migration 010 run?")
        print(msg)
        return {"status": "error", "error": msg}

    print(f"\nTrend Scout — {run_date.isoformat()}")
    print(f"  Sources:        {len(sources)}"
          + (f"  (offering_hint={offering})" if offering else ""))
    print(f"  Lookback:       {lookback_days} days")
    print(f"  Mode:           {'DRY RUN (no writes)' if dry_run else 'live'}")
    print()

    tracked = load_tracked_keywords()
    print(f"  Tracked keywords loaded: {len(tracked)} (for novelty checking)")
    warn_unhealthy_sources()
    print()

    # ---- Harvest ----------------------------------------------------------
    mentions, harvest_stats = harvest(sources, lookback_days, max_items, dry_run)
    print()
    print(f"  Sources polled:      {harvest_stats['sources_polled']} "
          f"({harvest_stats['sources_ok']} ok, {harvest_stats['sources_failed']} failed)")
    print(f"  Items fetched:       {harvest_stats['items_fetched']}")
    print(f"  New mentions:        {harvest_stats['mentions_new']}")
    print(f"  Already seen:        {harvest_stats['mentions_duplicate']}")
    if "mentions_in_window" in harvest_stats:
        print(f"  Scoring window:      {harvest_stats['mentions_in_window']} "
              f"mention(s) over {lookback_days} days")

    if not mentions:
        print("\n  No mentions in the scoring window — nothing to score.")
        if not dry_run:
            record_agent_run(AGENT_NAME, "success", 0, [],
                             round(time.monotonic() - started, 2),
                             {**harvest_stats, "run_date": run_date.isoformat(),
                              "reason": "no new mentions"})
        return {"status": "success", "candidates": 0, **harvest_stats}

    # ---- Extract ----------------------------------------------------------
    agg = extract_phrases(mentions)
    print(f"  Phrases extracted:   {len(agg):,}")

    survivors = filter_candidates(agg, tracked, min_mentions, min_spread)
    print(f"  After filtering:     {len(survivors)} "
          f"(>={min_mentions} mentions, >={min_spread} sources, commercially shaped)")

    if not survivors:
        print("\n  Nothing cleared the thresholds. Try --days 30 or --min-mentions 1.")
        if not dry_run:
            record_agent_run(AGENT_NAME, "success", 0, [],
                             round(time.monotonic() - started, 2),
                             {**harvest_stats, "run_date": run_date.isoformat(),
                              "phrases_extracted": len(agg), "reason": "no survivors"})
        return {"status": "success", "candidates": 0, **harvest_stats}

    # ---- Classify: rules first --------------------------------------------
    offerings = sorted({t["offering"] for t in tracked if t["offering"]})
    candidates: list[dict] = []
    unresolved: list[dict] = []

    for entry in survivors:
        rule_offering = entry.get("rule_offering")
        hint_offering = None
        if entry["offering_hints"]:
            top_hint, hint_support = max(entry["offering_hints"].items(),
                                         key=lambda kv: kv[1])
            # One offering-specific outlet mentioning a generic phrase proves
            # nothing — that's how "customer service" got labelled AS400 off a
            # single IT Jungle article. Require the hint to recur.
            if hint_support >= MIN_HINT_SUPPORT:
                hint_offering = top_hint

        base = {
            "phrase":          entry["phrase"],
            "keyword":         entry["phrase"],
            "display_keyword": entry["phrase"],
            "mention_count":   entry["mention_count"],
            "weighted_count":  entry["weighted_count"],
            "source_spread":   len(entry["source_ids"]),
            "category_spread": len(entry["categories"]),
            "first_seen":      entry["first_seen"],
            "last_seen":       entry["last_seen"],
            "evidence":        entry["evidence"],
            "has_commercial":  entry["has_commercial"],
            "nearest_tracked": entry["nearest_tracked"],
            "nearest_similarity": entry["nearest_similarity"],
            "is_novel":        entry["is_novel"],
            "intent":          _infer_intent(entry["phrase"], entry["has_commercial"]),
        }

        if rule_offering:
            base["offering"] = rule_offering
            base["offering_confidence"] = 0.75
            base["classification_method"] = "rule"
            candidates.append(base)
        elif hint_offering:
            # The source itself is offering-specific (a Salesforce blog, an
            # InsurTech subreddit) — weaker than a token match but real.
            base["offering"] = hint_offering
            base["offering_confidence"] = 0.50
            base["classification_method"] = "source_hint"
            candidates.append(base)
        else:
            unresolved.append(entry)

    print(f"  Classified by rule:  {sum(1 for c in candidates if c['classification_method'] == 'rule')}")
    print(f"  Classified by hint:  {sum(1 for c in candidates if c['classification_method'] == 'source_hint')}")
    print(f"  Unresolved:          {len(unresolved)}")

    # ---- Classify: Claude for the residue ---------------------------------
    llm_used = False
    if use_llm and unresolved:
        batch = unresolved[:MAX_LLM_CANDIDATES]
        print(f"  Sending {len(batch)} unresolved phrase(s) to Claude (cheap tier)...")
        llm_results = classify_with_llm(batch, offerings)
        llm_used = bool(llm_results)
        for entry in batch:
            got = llm_results.get(entry["phrase"])
            if not got:
                continue
            # The LLM may reshape the phrase, so novelty must be rechecked
            # against the rewritten keyword, not the raw phrase.
            nearest, similarity = _nearest_tracked(got["keyword"], tracked)
            candidates.append({
                "phrase":          entry["phrase"],
                "keyword":         got["keyword"],
                "display_keyword": got["keyword"],
                "offering":        got["offering"],
                "offering_confidence": got["confidence"],
                "classification_method": "llm",
                # LLM intent wins when present; fall back to the rule so a
                # partial LLM response can't leave the candidate unscored.
                "intent":          got["intent"] or _infer_intent(
                                       got["keyword"], entry["has_commercial"]),
                "mention_count":   entry["mention_count"],
                "weighted_count":  entry["weighted_count"],
                "source_spread":   len(entry["source_ids"]),
                "category_spread": len(entry["categories"]),
                "first_seen":      entry["first_seen"],
                "last_seen":       entry["last_seen"],
                "evidence":        entry["evidence"],
                "has_commercial":  entry["has_commercial"],
                "nearest_tracked": nearest,
                "nearest_similarity": similarity,
                "is_novel":        similarity < NOVELTY_SIMILARITY_THRESHOLD,
            })
        print(f"  Claude classified:   {len(llm_results)}")
    elif unresolved:
        print("  Skipping LLM classification (--no-llm)")

    # Reshaping can collide two phrases onto one keyword — keep the loudest.
    deduped: dict[str, dict] = {}
    for c in candidates:
        key = _normalize(c["keyword"])
        c["keyword"] = key
        prior = deduped.get(key)
        if prior is None or c["weighted_count"] > prior["weighted_count"]:
            deduped[key] = c

    before_collapse = len(deduped)
    candidates = collapse_subphrases(sorted(deduped.values(),
                                            key=lambda c: -c["weighted_count"]))
    print(f"  Candidates:          {len(candidates)} "
          f"(deduped {before_collapse}, collapsed {before_collapse - len(candidates)} sub-phrases)")

    # ---- Semantic novelty re-check ----------------------------------------
    # Jaccard already ran as a cheap prefilter during extraction. It cannot see
    # paraphrase, so re-check the survivors — a small batch by this point —
    # against embeddings. This is the guard on recurring spend: every promoted
    # keyword costs money on every future rank-tracker run, and a paraphrase of
    # something already tracked buys nothing.
    semantic_stats = refine_novelty_semantic(candidates, tracked)
    if semantic_stats["checked"]:
        print(f"  Semantic novelty:    {semantic_stats['checked']} checked, "
              f"{semantic_stats['demoted']} reclassified as near-duplicates")
    elif semantic_stats["reason"]:
        print(f"  Semantic novelty:    skipped ({semantic_stats['reason']})")

    # ---- Volume -----------------------------------------------------------
    volume_stats = {"looked_up": 0, "matched": 0, "cost_usd": 0.0, "error": None}
    if use_volume and not dry_run:
        print(f"\n  Fetching Google Ads Keyword Planner volume "
              f"(up to {max_volume_lookups} keywords)...")
        volume_stats = enrich_with_volume(candidates, max_volume_lookups)
        print(f"  Volume looked up:    {volume_stats['looked_up']} "
              f"({volume_stats['matched']} matched)  ~${volume_stats['cost_usd']:.2f}")
        if volume_stats["error"]:
            errors.append(f"volume lookup: {volume_stats['error']}")
    elif use_volume and dry_run:
        print("\n  Skipping volume lookup (dry run — it costs money)")
    else:
        print("\n  Skipping volume lookup (--no-volume)")

    # ---- Score ------------------------------------------------------------
    for c in candidates:
        score_candidate(c)
    candidates.sort(key=lambda c: -c["trend_score"])

    # ---- Write ------------------------------------------------------------
    write_stats = {"inserted": 0, "updated": 0}
    if not dry_run:
        write_stats = upsert_candidates(candidates)
        print(f"\n  Candidates written:  {write_stats['inserted']} new, "
              f"{write_stats['updated']} updated")
        # Attach DB ids so the Excel export is actionable (--promote --ids).
        id_rows = fetch_all(
            "SELECT id, candidate_keyword FROM keyword_candidates WHERE candidate_keyword = ANY(%s)",
            ([c["keyword"] for c in candidates],),
        )
        id_by_kw = {r["candidate_keyword"]: r["id"] for r in id_rows}
        for c in candidates:
            c["db_id"] = id_by_kw.get(c["keyword"])
    else:
        print(f"\n  DRY RUN — would write {len(candidates)} candidate(s)")

    duration = time.monotonic() - started
    stats = {
        "run_date":          run_date.isoformat(),
        "lookback_days":     lookback_days,
        "offering_filter":   offering,
        **harvest_stats,
        "phrases_extracted": len(agg),
        "candidates":        len(candidates),
        "llm_used":          llm_used,
        "volume_looked_up":  volume_stats["looked_up"],
        "volume_matched":    volume_stats["matched"],
        "volume_cost_usd":   volume_stats["cost_usd"],
        "inserted":          write_stats["inserted"],
        "updated":           write_stats["updated"],
    }

    # ---- Notify -----------------------------------------------------------
    top = candidates[:25]
    print_candidate_table(
        [{
            "id": c.get("db_id") or "—",
            "trend_score": c["trend_score"],
            "search_volume": c.get("search_volume"),
            "momentum_ratio": c.get("momentum_ratio"),
            "mention_count": c["mention_count"],
            "suggested_offering": c.get("offering"),
            "display_keyword": c.get("display_keyword"),
        } for c in top],
        f"TOP {len(top)} EMERGING KEYWORDS",
    )

    reports: dict[str, Path] = {}
    if not dry_run and not skip_reports:
        reports = write_reports(candidates, run_date, stats)
        print()
        for kind, path in reports.items():
            print(f"  {kind.title()} report: {path}")

    print()
    print(f"  Duration:            {duration:.1f}s")
    print(f"  Estimated cost:      ${volume_stats['cost_usd']:.2f} "
          f"(Keyword Planner) + LLM classification")
    print()
    print("  Next: review with --list-candidates, approve in SQL, then --promote.")
    print("  Nothing has been added to the tracked keyword set.")
    print()

    if not dry_run:
        record_agent_run(
            AGENT_NAME,
            status="success" if not errors else "partial",
            records_processed=len(candidates),
            errors=errors,
            duration_seconds=round(duration, 2),
            metadata={**stats, "reports": {k: str(v) for k, v in reports.items()}},
        )

    return {"status": "success" if not errors else "partial", **stats}


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description=f"{profile().brand_name} Trend Scout — discover emerging "
                    f"industry keywords",
    )
    parser.add_argument("--days", type=int, default=DEFAULT_LOOKBACK_DAYS,
                        help=f"Lookback window in days (default: {DEFAULT_LOOKBACK_DAYS})")
    parser.add_argument("--offering",
                        help="Only poll sources hinted at this offering (e.g. 'AI')")
    parser.add_argument("--max-items", type=int, default=DEFAULT_MAX_ITEMS_PER_SOURCE,
                        help=f"Max items per source (default: {DEFAULT_MAX_ITEMS_PER_SOURCE})")
    parser.add_argument("--min-mentions", type=int, default=MIN_MENTIONS,
                        help=f"Minimum mentions for a phrase (default: {MIN_MENTIONS})")
    parser.add_argument("--min-spread", type=int, default=MIN_SOURCE_SPREAD,
                        help=f"Minimum distinct sources (default: {MIN_SOURCE_SPREAD})")
    parser.add_argument("--max-volume-lookups", type=int, default=DEFAULT_MAX_VOLUME_LOOKUPS,
                        help=f"Cap Keyword Planner lookups (default: {DEFAULT_MAX_VOLUME_LOOKUPS})")
    parser.add_argument("--no-llm", dest="use_llm", action="store_false",
                        help="Skip Claude classification; token rules only")
    parser.add_argument("--no-volume", dest="use_volume", action="store_false",
                        help="Skip the paid Keyword Planner lookup")
    parser.add_argument("--dry-run", action="store_true",
                        help="Harvest and score but write nothing (also skips paid lookups)")
    parser.add_argument("--skip-reports", action="store_true",
                        help="Skip Excel/markdown generation")

    parser.add_argument("--list-candidates", action="store_true",
                        help="List existing candidates instead of running discovery")
    parser.add_argument("--status", default="new",
                        help="With --list-candidates: filter by status, or 'all' (default: new)")
    parser.add_argument("--min-score", type=float, default=0.0,
                        help="Minimum trend_score for --list-candidates / --promote")
    parser.add_argument("--limit", type=int, default=50,
                        help="With --list-candidates: row cap (default: 50)")

    parser.add_argument("--promote", action="store_true",
                        help="Human gate: move candidates into the tracked keyword set")
    parser.add_argument("--ids", help="With --promote: comma-separated candidate IDs")

    parser.add_argument("--verbose", "-v", action="store_true", help="Debug logging")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(name)s  %(levelname)s  %(message)s",
    )

    if args.list_candidates:
        rows = list_candidates(min_score=args.min_score, offering=args.offering,
                               status=args.status, limit=args.limit)
        print_candidate_table(
            rows, f"KEYWORD CANDIDATES (status={args.status}, score>={args.min_score})")
        print()
        return

    if args.promote:
        ids = None
        if args.ids:
            ids = [int(x) for x in args.ids.replace(" ", "").split(",") if x]
        if not ids and args.min_score <= 0:
            parser.error(
                "--promote needs --ids, or --min-score with candidates already "
                "marked status='approved'. Refusing to bulk-promote unreviewed "
                "candidates into the paid tracking set."
            )
        result = promote_candidates(candidate_ids=ids,
                                    min_score=args.min_score if not ids else None,
                                    offering=args.offering,
                                    dry_run=args.dry_run)
        print()
        print(f"  Promoted:    {result['promoted']}")
        print(f"  Duplicates:  {result['duplicates']} (already in keywords)")
        print(f"  Skipped:     {result['skipped']}")
        for d in result["details"]:
            print(f"    [{d['action']:<14}] {_safe_console(d['keyword'])[:60]}"
                  f"  ({d.get('offering', '?')})")
        if result["promoted"] and not args.dry_run:
            print()
            print("  Next: rank-track the new keywords —")
            print("    python -m keyword_intelligence.rank_tracker --offering "
                  "\"<offering>\" --all")
        print()
        return

    run(
        lookback_days=args.days,
        offering=args.offering,
        max_items=args.max_items,
        min_mentions=args.min_mentions,
        min_spread=args.min_spread,
        max_volume_lookups=args.max_volume_lookups,
        use_llm=args.use_llm,
        use_volume=args.use_volume,
        dry_run=args.dry_run,
        skip_reports=args.skip_reports,
    )


if __name__ == "__main__":
    main()
