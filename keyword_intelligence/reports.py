"""
Ranking Report Generator
========================

Generates Excel reports from keyword_rankings data. Designed to match the
format SEO executives are already used to (the wide-format Excel with dates
as columns), while adding structured analysis sheets.

Usage
-----
    # Generate report for the latest run
    python -m keyword_intelligence.reports

    # Generate report for a specific date range
    python -m keyword_intelligence.reports --start 2026-02-01 --end 2026-04-15

    # Filter by offering
    python -m keyword_intelligence.reports --offering "Insurance Broker Software"
"""

from __future__ import annotations

import argparse
import logging
import sys
from datetime import date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common.config import settings
from common.database import fetch_all


logger = logging.getLogger("reports")

# Styles
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D5E8F0", end_color="D5E8F0", fill_type="solid")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
STRIKING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _style_header(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_cell(ws, row: int, col: int) -> None:
    cell = ws.cell(row=row, column=col)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_ranking_data(start_date: date | None, end_date: date | None,
                      offering: str | None = None) -> dict:
    """
    Load ranking data grouped by keyword with all available date snapshots.
    Returns a dict with structure needed for each report sheet.
    """
    # Get all unique dates for which we have data
    date_sql = """
        SELECT DISTINCT kr.date
        FROM keyword_rankings kr
        JOIN keywords k ON k.id = kr.keyword_id
        WHERE k.status = 'active'
    """
    params: list = []
    if start_date:
        date_sql += " AND kr.date >= %s"
        params.append(start_date)
    if end_date:
        date_sql += " AND kr.date <= %s"
        params.append(end_date)
    if offering:
        date_sql += " AND k.offering = %s"
        params.append(offering)
    date_sql += " ORDER BY kr.date"

    date_rows = fetch_all(date_sql, params)
    dates = [r["date"] for r in date_rows]

    if not dates:
        return {"dates": [], "keywords": [], "by_keyword": {}}

    # Get all rankings (both DataForSEO and GSC)
    sql = """
        SELECT k.keyword, k.offering, k.target_url,
               kr.date, kr.rank_position, kr.rank_bucket, kr.url_found,
               kr.source, kr.clicks, kr.impressions, kr.ctr
        FROM keyword_rankings kr
        JOIN keywords k ON k.id = kr.keyword_id
        WHERE k.status = 'active'
    """
    params2: list = []
    if start_date:
        sql += " AND kr.date >= %s"
        params2.append(start_date)
    if end_date:
        sql += " AND kr.date <= %s"
        params2.append(end_date)
    if offering:
        sql += " AND k.offering = %s"
        params2.append(offering)
    sql += " ORDER BY k.offering, k.keyword, kr.date"

    rows = fetch_all(sql, params2)

    # Group by keyword, separating DataForSEO rankings from GSC metrics
    by_keyword: dict[str, dict] = {}
    for r in rows:
        kw = r["keyword"]
        if kw not in by_keyword:
            by_keyword[kw] = {
                "offering": r["offering"],
                "target_url": r["target_url"],
                "rankings": {},   # keyed by date — DataForSEO/manual data
                "gsc": {},        # keyed by date — GSC data
            }
        source = r["source"]
        entry = {
            "position": r["rank_position"],
            "bucket": r["rank_bucket"],
            "url_found": r["url_found"],
            "clicks": r.get("clicks"),
            "impressions": r.get("impressions"),
            "ctr": r.get("ctr"),
        }
        if source == "gsc":
            by_keyword[kw]["gsc"][r["date"]] = entry
        else:
            by_keyword[kw]["rankings"][r["date"]] = entry

    keywords_sorted = sorted(by_keyword.keys(), key=lambda k: (by_keyword[k]["offering"] or "", k))

    return {
        "dates": dates,
        "keywords": keywords_sorted,
        "by_keyword": by_keyword,
    }


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_summary_sheet(wb: Workbook, data: dict) -> None:
    """Sheet 1: Bucket distribution per date snapshot."""
    ws = wb.active
    ws.title = "Summary"

    dates = data["dates"]
    by_keyword = data["by_keyword"]

    # Title
    ws.cell(row=1, column=1, value="Damco Keyword Rank Tracker — Summary")
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(dates))

    # Bucket summary table
    bucket_order = ["1-5", "5-10", "10-20", "20-50", "50+", "not-found"]
    start_row = 3
    headers = ["Rank Bucket"] + [d.isoformat() if isinstance(d, date) else str(d) for d in dates]
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    _style_header(ws, start_row, len(headers))

    for i, bucket_name in enumerate(bucket_order):
        row = start_row + 1 + i
        ws.cell(row=row, column=1, value=bucket_name)
        ws.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=row, column=1).border = THIN_BORDER
        for j, d in enumerate(dates, 2):
            count = sum(
                1 for kw_data in by_keyword.values()
                if kw_data["rankings"].get(d, {}).get("bucket") == bucket_name
            )
            ws.cell(row=row, column=j, value=count)
            _style_data_cell(ws, row, j)

    # Total row
    total_row = start_row + 1 + len(bucket_order)
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=total_row, column=1).border = THIN_BORDER
    for j, d in enumerate(dates, 2):
        count = sum(1 for kw_data in by_keyword.values() if d in kw_data["rankings"])
        ws.cell(row=total_row, column=j, value=count)
        _style_data_cell(ws, total_row, j)

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def build_detailed_sheet(wb: Workbook, data: dict) -> None:
    """Sheet 2: Wide-format table with SERP rank + GSC metrics side by side."""
    ws = wb.create_sheet("Detailed Rankings")

    dates = data["dates"]
    keywords = data["keywords"]
    by_keyword = data["by_keyword"]

    # Check if any GSC data exists
    has_gsc = any(kw_data["gsc"] for kw_data in by_keyword.values())

    # Headers: Keyword | Offering | date1 | date2 | ... | GSC Avg Pos | GSC Clicks | GSC Impr | GSC CTR
    headers = ["Keyword", "Offering"] + [
        d.isoformat() if isinstance(d, date) else str(d) for d in dates
    ]
    if has_gsc:
        headers += ["GSC Avg Pos (14d)", "GSC Clicks", "GSC Impressions", "GSC CTR"]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    gsc_start_col = 3 + len(dates)  # first GSC column

    for i, kw in enumerate(keywords, 2):
        kw_data = by_keyword[kw]
        ws.cell(row=i, column=1, value=kw)
        ws.cell(row=i, column=1).border = THIN_BORDER
        ws.cell(row=i, column=2, value=kw_data["offering"])
        ws.cell(row=i, column=2).border = THIN_BORDER

        # DataForSEO SERP positions
        for j, d in enumerate(dates, 3):
            ranking = kw_data["rankings"].get(d)
            if ranking:
                pos = ranking["position"]
                val = pos if pos is not None else "N/F"
            else:
                val = ""
            ws.cell(row=i, column=j, value=val)
            _style_data_cell(ws, i, j)

            # Color-code positions
            if isinstance(val, int):
                cell = ws.cell(row=i, column=j)
                if val <= 10:
                    cell.fill = GOOD_FILL
                elif val <= 20:
                    cell.fill = STRIKING_FILL

        # GSC metrics — use the latest GSC date available for this keyword
        if has_gsc:
            gsc_dates = sorted(kw_data["gsc"].keys())
            if gsc_dates:
                latest_gsc = kw_data["gsc"][gsc_dates[-1]]
                gsc_pos = latest_gsc["position"]
                ws.cell(row=i, column=gsc_start_col, value=gsc_pos if gsc_pos is not None else "N/A")
                _style_data_cell(ws, i, gsc_start_col)
                if isinstance(gsc_pos, (int, float)) and gsc_pos <= 20:
                    ws.cell(row=i, column=gsc_start_col).fill = STRIKING_FILL if gsc_pos > 10 else GOOD_FILL

                ws.cell(row=i, column=gsc_start_col + 1, value=latest_gsc.get("clicks") or 0)
                _style_data_cell(ws, i, gsc_start_col + 1)

                ws.cell(row=i, column=gsc_start_col + 2, value=latest_gsc.get("impressions") or 0)
                _style_data_cell(ws, i, gsc_start_col + 2)

                ctr = latest_gsc.get("ctr")
                if ctr is not None:
                    ws.cell(row=i, column=gsc_start_col + 3, value=ctr)
                    ws.cell(row=i, column=gsc_start_col + 3).number_format = "0.00%"
                _style_data_cell(ws, i, gsc_start_col + 3)
            else:
                for offset in range(4):
                    ws.cell(row=i, column=gsc_start_col + offset, value="-")
                    _style_data_cell(ws, i, gsc_start_col + offset)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25
    for col in range(3, 3 + len(dates)):
        ws.column_dimensions[get_column_letter(col)].width = 14
    if has_gsc:
        ws.column_dimensions[get_column_letter(gsc_start_col)].width = 18
        ws.column_dimensions[get_column_letter(gsc_start_col + 1)].width = 12
        ws.column_dimensions[get_column_letter(gsc_start_col + 2)].width = 14
        ws.column_dimensions[get_column_letter(gsc_start_col + 3)].width = 10


def build_movement_sheet(wb: Workbook, data: dict) -> None:
    """Sheet 3: Movement between the two most recent snapshots."""
    ws = wb.create_sheet("Movement")

    dates = data["dates"]
    by_keyword = data["by_keyword"]

    if len(dates) < 2:
        ws.cell(row=1, column=1, value="Need at least 2 date snapshots to compute movement.")
        return

    prev_date = dates[-2]
    curr_date = dates[-1]

    headers = ["Keyword", "Offering", f"Prev ({prev_date})", f"Current ({curr_date})",
               "Change", "Direction"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    movements: list[tuple[str, str | None, int | None, int | None, int | None]] = []
    for kw, kw_data in by_keyword.items():
        prev_rank = kw_data["rankings"].get(prev_date, {}).get("position")
        curr_rank = kw_data["rankings"].get(curr_date, {}).get("position")

        if prev_rank is not None and curr_rank is not None:
            change = prev_rank - curr_rank  # positive = improved
        else:
            change = None
        movements.append((kw, kw_data["offering"], prev_rank, curr_rank, change))

    # Sort: biggest improvements first, then declines, then unchanged
    movements.sort(key=lambda m: (m[4] is None, -(m[4] or 0)))

    for i, (kw, offering, prev_r, curr_r, change) in enumerate(movements, 2):
        ws.cell(row=i, column=1, value=kw).border = THIN_BORDER
        ws.cell(row=i, column=2, value=offering).border = THIN_BORDER
        ws.cell(row=i, column=3, value=prev_r if prev_r is not None else "N/F")
        _style_data_cell(ws, i, 3)
        ws.cell(row=i, column=4, value=curr_r if curr_r is not None else "N/F")
        _style_data_cell(ws, i, 4)

        if change is not None:
            ws.cell(row=i, column=5, value=change)
            direction = "Improved" if change > 0 else ("Declined" if change < 0 else "Stable")
            ws.cell(row=i, column=6, value=direction)
            cell5 = ws.cell(row=i, column=5)
            cell6 = ws.cell(row=i, column=6)
            if change > 0:
                cell5.fill = GOOD_FILL
                cell6.fill = GOOD_FILL
            elif change < 0:
                cell5.fill = BAD_FILL
                cell6.fill = BAD_FILL
        else:
            ws.cell(row=i, column=5, value="-")
            ws.cell(row=i, column=6, value="N/A")

        _style_data_cell(ws, i, 5)
        _style_data_cell(ws, i, 6)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25
    for col in range(3, 7):
        ws.column_dimensions[get_column_letter(col)].width = 16


def build_striking_distance_sheet(wb: Workbook, data: dict) -> None:
    """Sheet 4: Keywords at positions 11-20 in the latest snapshot."""
    ws = wb.create_sheet("Striking Distance")

    dates = data["dates"]
    by_keyword = data["by_keyword"]

    if not dates:
        ws.cell(row=1, column=1, value="No data available.")
        return

    latest_date = dates[-1]

    headers = ["Keyword", "Offering", "Position", "URL Found", "Target URL"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    striking: list[tuple[str, dict, int, str | None]] = []
    for kw, kw_data in by_keyword.items():
        ranking = kw_data["rankings"].get(latest_date)
        if ranking and ranking["position"] and 11 <= ranking["position"] <= 20:
            striking.append((kw, kw_data, ranking["position"], ranking.get("url_found")))

    striking.sort(key=lambda s: s[2])

    if not striking:
        ws.cell(row=2, column=1, value="No keywords in striking distance (positions 11-20).")
        return

    for i, (kw, kw_data, pos, url_found) in enumerate(striking, 2):
        ws.cell(row=i, column=1, value=kw).border = THIN_BORDER
        ws.cell(row=i, column=2, value=kw_data["offering"]).border = THIN_BORDER
        ws.cell(row=i, column=3, value=pos)
        ws.cell(row=i, column=3).fill = STRIKING_FILL
        _style_data_cell(ws, i, 3)
        ws.cell(row=i, column=4, value=url_found or "").border = THIN_BORDER
        ws.cell(row=i, column=5, value=kw_data["target_url"] or "").border = THIN_BORDER

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 50


def build_gsc_sheet(wb: Workbook, data: dict) -> None:
    """Sheet 5: GSC Performance — dedicated view of Google's own metrics."""
    ws = wb.create_sheet("GSC Performance")

    by_keyword = data["by_keyword"]

    # Collect all keywords that have GSC data
    gsc_rows: list[tuple[str, dict, dict]] = []
    for kw, kw_data in by_keyword.items():
        gsc_dates = sorted(kw_data.get("gsc", {}).keys())
        if gsc_dates:
            latest_gsc = kw_data["gsc"][gsc_dates[-1]]
            gsc_rows.append((kw, kw_data, latest_gsc))

    if not gsc_rows:
        ws.cell(row=1, column=1, value="No GSC data available. Run GSC enrichment first:")
        ws.cell(row=2, column=1, value="  python -m keyword_intelligence.gsc_enrichment")
        ws.cell(row=3, column=1)
        ws.cell(row=4, column=1, value="Or run the rank tracker with GSC enabled (default):")
        ws.cell(row=5, column=1, value="  python -m keyword_intelligence.rank_tracker")
        return

    headers = ["Keyword", "Offering", "SERP Rank", "GSC Avg Pos", "Clicks (14d)",
               "Impressions (14d)", "CTR", "Gap (SERP vs GSC)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header(ws, 1, len(headers))

    # Sort by impressions descending (most visible keywords first)
    gsc_rows.sort(key=lambda x: x[2].get("impressions") or 0, reverse=True)

    for i, (kw, kw_data, gsc) in enumerate(gsc_rows, 2):
        # Get latest SERP rank for comparison
        ranking_dates = sorted(kw_data.get("rankings", {}).keys())
        serp_pos = None
        if ranking_dates:
            serp_pos = kw_data["rankings"][ranking_dates[-1]].get("position")

        gsc_pos = gsc.get("position")
        clicks = gsc.get("clicks") or 0
        impressions = gsc.get("impressions") or 0
        ctr = gsc.get("ctr")

        # Gap: positive = SERP rank is better than GSC avg (good), negative = worse
        gap = None
        if serp_pos is not None and gsc_pos is not None:
            gap = round(gsc_pos) - serp_pos  # positive means GSC sees you worse

        ws.cell(row=i, column=1, value=kw).border = THIN_BORDER
        ws.cell(row=i, column=2, value=kw_data["offering"]).border = THIN_BORDER

        ws.cell(row=i, column=3, value=serp_pos if serp_pos is not None else "N/A")
        _style_data_cell(ws, i, 3)

        ws.cell(row=i, column=4, value=round(gsc_pos) if gsc_pos is not None else "N/A")
        _style_data_cell(ws, i, 4)

        ws.cell(row=i, column=5, value=clicks)
        _style_data_cell(ws, i, 5)

        ws.cell(row=i, column=6, value=impressions)
        _style_data_cell(ws, i, 6)

        if ctr is not None:
            ws.cell(row=i, column=7, value=ctr)
            ws.cell(row=i, column=7).number_format = "0.00%"
        _style_data_cell(ws, i, 7)

        if gap is not None:
            ws.cell(row=i, column=8, value=gap)
            cell = ws.cell(row=i, column=8)
            if gap > 3:
                cell.fill = BAD_FILL   # GSC sees much worse than SERP snapshot
            elif gap < -3:
                cell.fill = GOOD_FILL  # GSC sees better (trending up)
        _style_data_cell(ws, i, 8)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25
    for col in range(3, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generate_report(
    start_date: date | None = None,
    end_date: date | None = None,
    offering: str | None = None,
    output_path: str | None = None,
) -> Path:
    """Generate an Excel ranking report and return the file path."""
    data = load_ranking_data(start_date, end_date, offering)

    if not data["dates"]:
        logger.warning("No ranking data found for the given parameters")
        print("No ranking data found. Run the rank tracker first.")
        sys.exit(1)

    wb = Workbook()
    build_summary_sheet(wb, data)
    build_detailed_sheet(wb, data)
    build_movement_sheet(wb, data)
    build_striking_distance_sheet(wb, data)
    build_gsc_sheet(wb, data)

    # Output path
    if output_path:
        path = Path(output_path)
    else:
        reports_dir = settings.OUTPUTS_DIR / "reports"
        reports_dir.mkdir(parents=True, exist_ok=True)
        latest = data["dates"][-1]
        date_str = latest.isoformat() if isinstance(latest, date) else str(latest)
        suffix = f"_{offering.replace(' ', '_')}" if offering else ""
        path = reports_dir / f"ranking_report_{date_str}{suffix}.xlsx"

    wb.save(str(path))
    print(f"\n  Report saved to: {path}")
    print(f"  Sheets: Summary, Detailed Rankings, Movement, Striking Distance, GSC Performance")
    print(f"  Keywords: {len(data['keywords'])}")
    print(f"  Date range: {data['dates'][0]} to {data['dates'][-1]}")
    print()

    return path


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate keyword ranking Excel report")
    parser.add_argument("--start", help="Start date (YYYY-MM-DD)")
    parser.add_argument("--end", help="End date (YYYY-MM-DD)")
    parser.add_argument("--offering", help="Filter by offering")
    parser.add_argument("--output", "-o", help="Output file path")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(name)s  %(message)s")

    start = date.fromisoformat(args.start) if args.start else None
    end = date.fromisoformat(args.end) if args.end else None

    generate_report(start_date=start, end_date=end, offering=args.offering, output_path=args.output)


if __name__ == "__main__":
    main()
